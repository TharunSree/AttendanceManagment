# In academics/views.py
import calendar
import csv
import json
import os
from datetime import datetime
from itertools import chain

from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.exceptions import PermissionDenied, ObjectDoesNotExist
from django.core.management import call_command
from django.db import transaction
from django.db.models import Count, Sum
from django.db.models import Q
from django.db.models.functions import TruncMonth
from django.forms import inlineformset_factory, formset_factory
from django.http import Http404, JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pyppeteer import launch


from academics.forms import EditStudentForm, AttendanceSettingsForm, TimeSlotForm, MarkAttendanceForm, \
    TimetableEntryForm, SubstitutionForm, AttendanceReportForm, AnnouncementForm, CriterionFormSet, MarkingSchemeForm, \
    MarkSelectForm, BulkMarksImportForm, MarksReportForm, ExtraClassForm, SmtpSettingsForm, BulkEmailForm, \
    AcademicSessionForm, AcademicSessionModelForm, SupplementaryMarkForm
from accounts.decorators import nav_item
from accounts.models import Profile, UserActivityLog
from .email_utils import send_database_email
from .forms import StudentGroupForm, CourseForm, AddStudentForm, SubjectForm
from .models import StudentGroup, AttendanceSettings, Course, Subject, \
    Timetable, AttendanceRecord, CourseSubject, TimeSlot, ClassCancellation, DailySubstitution, Announcement, \
    UserNotificationStatus, MarkingScheme, Mark, Criterion, ExtraClass, AcademicSession, ResultPublication, \
    StudentSubjectStatus
import subprocess
import logging

logger = logging.getLogger(__name__)

# ... other views ...

@login_required
@permission_required('academics.view_attendance_settings', raise_exception=True)
@nav_item(title="Attendance Settings", icon="iconsminds-gears", url_name="academics:admin_settings",
          permission='academics.view_attendance_settings', group='application_settings', order=1)
def admin_settings_view(request):
    # Load settings and current session initially for context
    settings_obj = AttendanceSettings.load()
    try:
        current_session = AcademicSession.objects.get(is_current=True)
    except AcademicSession.DoesNotExist:
        current_session = None

    if request.method == 'POST':
        # Check which form was submitted and handle it
        if 'submit_timeslot' in request.POST:
            timeslot_form = TimeSlotForm(request.POST)
            if timeslot_form.is_valid():
                timeslot_form.save()
                messages.success(request, "New time slot has been created.")
                return redirect('academics:admin_settings')  # Redirect on success

        elif 'submit_settings' in request.POST:
            settings_form = AttendanceSettingsForm(request.POST, instance=settings_obj)
            if settings_form.is_valid():
                settings_form.save()
                messages.success(request, "General system settings have been updated.")
                return redirect('academics:admin_settings')  # Redirect on success

        elif 'set_academic_session' in request.POST:
            session_set_form = AcademicSessionForm(request.POST)
            if session_set_form.is_valid():
                new_current_session = session_set_form.cleaned_data['current_session']
                AcademicSession.objects.exclude(pk=new_current_session.pk).update(is_current=False)
                new_current_session.is_current = True
                new_current_session.save()
                messages.success(request, f"'{new_current_session.name}' is now the current academic session.")
                return redirect('academics:admin_settings')  # Redirect on success

        elif 'create_academic_session' in request.POST:
            academic_session_create_form = AcademicSessionModelForm(request.POST)
            if academic_session_create_form.is_valid():
                academic_session_create_form.save()
                messages.success(request, "New academic session has been created.")
                return redirect('academics:admin_settings')  # Redirect on success
    
    # If it's a GET request or a POST with errors, initialize all forms
    timeslot_form = TimeSlotForm()
    settings_form = AttendanceSettingsForm(instance=settings_obj)
    session_set_form = AcademicSessionForm(initial={'current_session': current_session})
    academic_session_create_form = AcademicSessionModelForm()
    
    # If a specific form was submitted via POST and was invalid,
    # that form instance (containing errors) will be used in the context.
    # We must ensure the context gets the form that may have been created in the POST block.
    # A simple way is to check the POST data again.
    if request.method == 'POST':
        if 'submit_timeslot' in request.POST:
            timeslot_form = TimeSlotForm(request.POST)
        elif 'submit_settings' in request.POST:
            settings_form = AttendanceSettingsForm(request.POST, instance=settings_obj)
        elif 'set_academic_session' in request.POST:
            session_set_form = AcademicSessionForm(request.POST)
        elif 'create_academic_session' in request.POST:
            academic_session_create_form = AcademicSessionModelForm(request.POST)


    all_timeslots = TimeSlot.objects.all()
    all_sessions = AcademicSession.objects.all().order_by('-start_year')

    context = {
        'timeslots': all_timeslots,
        'timeslot_form': timeslot_form,
        'settings_form': settings_form,
        'session_form': session_set_form,
        'academic_session_create_form': academic_session_create_form,
        'academic_sessions': all_sessions,
        'page_title': 'Application Settings'
    }
    return render(request, 'academics/admin_settings.html', context)

@login_required
@permission_required('academics.delete_academicsession', raise_exception=True)
def academic_session_delete_view(request, pk):
    session = get_object_or_404(AcademicSession, pk=pk)
    if request.method == 'POST':
        if session.is_current:
            messages.error(request, "Cannot delete the currently active academic session.")
        else:
            session.delete()
            messages.success(request, "Academic session has been deleted.")
    return redirect('academics:admin_settings')


# In academics/views.py

@login_required
@permission_required('auth.view_user', raise_exception=True)
def admin_student_list_view(request, group_id):
    student_group = get_object_or_404(StudentGroup, pk=group_id)
    students = User.objects.filter(
        profile__student_group=student_group, profile__role='student'
    ).select_related('profile').order_by('first_name', 'last_name')

    settings = AttendanceSettings.load()
    required_percentage = settings.required_percentage

    # Determine the latest semester to calculate attendance against
    latest_semester_num = CourseSubject.objects.filter(
        course=student_group.course
    ).order_by('-semester').values_list('semester', flat=True).first()

    students_with_attendance = []
    if latest_semester_num:
        for student in students:
            # This logic calculates the attendance percentage for each student
            subjects_for_semester = CourseSubject.objects.filter(
                course=student_group.course, semester=latest_semester_num
            )
            total_classes = AttendanceRecord.objects.filter(
                timetable__student_group=student_group,
                timetable__subject__in=subjects_for_semester
            ).values('date', 'timetable_id').distinct().count()

            student_records = AttendanceRecord.objects.filter(
                student=student, timetable__subject__in=subjects_for_semester
            )
            present_count = student_records.filter(status='Present').count()

            percentage = (present_count / total_classes * 100) if total_classes > 0 else None

            students_with_attendance.append({
                'student': student,
                'attendance_percentage': percentage
            })
    else:
        # If no semester, just list students without attendance data
        students_with_attendance = [{'student': s, 'attendance_percentage': None} for s in students]

    # --- FIX STARTS HERE ---

    # 1. Check the GET parameter from the form submission
    show_low_attendance_only = request.GET.get('low_attendance_filter') == 'on'

    # 2. If the filter is active, rebuild the list with only matching students
    if show_low_attendance_only:
        students_with_attendance = [
            item for item in students_with_attendance
            if item['attendance_percentage'] is not None and item['attendance_percentage'] < required_percentage
        ]

    # --- FIX ENDS HERE ---

    context = {
        'student_group': student_group,
        'students_with_attendance': students_with_attendance,
        'total_students': students.count(),
        'latest_semester_num': latest_semester_num,
        'required_percentage': required_percentage,
        # 3. Pass the flag back to the template to keep the checkbox state
        'show_low_attendance_only': show_low_attendance_only,
    }
    return render(request, 'academics/admin_student_list.html', context)


@login_required
@permission_required('academics.view_studentgroup', raise_exception=True)
@nav_item(title="Manage Classes", icon="iconsminds-network", url_name="academics:admin_select_class",
          permission='academics.view_studentgroup', group='admin_management', order=30)
def admin_select_class_view(request):
    """
    This view fetches all existing classes (StudentGroup) from the database
    and passes them to the template, so the admin can choose one.
    """
    # .select_related() is a performance optimization that fetches the
    # related course and session in the same database query.
    student_groups = StudentGroup.objects.all().select_related('course')

    context = {
        'student_groups': student_groups,
    }
    return render(request, 'academics/admin_select_class.html', context)


@login_required
@permission_required('academics.view_attendancerecord')  # Or a more general permission
def admin_student_attendance_detail_view(request, student_id):
    student = get_object_or_404(User, pk=student_id, profile__role='student')
    student_group = student.profile.student_group if hasattr(student, 'profile') else None

    # --- Get view type and filters from request ---
    view_type = request.GET.get('view_type', 'semester')
    selected_semester = request.GET.get('semester')
    selected_month_str = request.GET.get('month')
    selected_date_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))

    # --- Initialize context dictionary ---
    context = {
        'student': student, 'student_group': student_group, 'view_type': view_type,
        'available_semesters': [], 'selected_semester': None,
        'subject_attendance_data': [], 'subject_attendance_data_json': '[]',
        'overall_attended_sem': 0, 'overall_absent_sem': 0, 'overall_percentage_sem': 0,
        'monthly_subject_data': [], 'monthly_subject_data_json': '[]',
        'available_months': [], 'selected_month': selected_month_str,
        'overall_attended_month': 0, 'overall_absent_month': 0, 'overall_percentage_month': 0,
        'daily_attendance_data': [], 'selected_date': selected_date_str,
        'marks_data': {}
    }

    if student_group and student_group.course:
        all_course_subjects_qs = CourseSubject.objects.filter(course=student_group.course)
        context['available_semesters'] = all_course_subjects_qs.values_list('semester', flat=True).distinct().order_by(
            'semester')

        if not selected_semester and context['available_semesters']:
            selected_semester = context['available_semesters'].last()
        context['selected_semester'] = int(selected_semester) if selected_semester and str(
            selected_semester).isdigit() else None

        if context['selected_semester']:
            # This base queryset gets ALL of the student's attendance records for the semester,
            # regardless of whether they are from a regular class or an extra class.
            student_records_sem = AttendanceRecord.objects.filter(
                student=student
            ).filter(
                Q(timetable__subject__semester=context['selected_semester']) |
                Q(extra_class__subject__semester=context['selected_semester'])
            )

            if view_type == 'semester':
                subjects_for_semester = all_course_subjects_qs.filter(semester=context['selected_semester'])
                for cs in subjects_for_semester:
                    # Calculate total conducted classes by combining regular and extra classes.
                    # This counts distinct sessions for which attendance was marked.
                    regular_classes_count = AttendanceRecord.objects.filter(
                        timetable__student_group=student_group, timetable__subject=cs
                    ).values('date', 'timetable__time_slot').distinct().count()

                    extra_classes_count = ExtraClass.objects.filter(
                        class_group=student_group, subject=cs
                    ).count()

                    total_classes = regular_classes_count + extra_classes_count

                    if total_classes > 0:
                        # Count attended classes from the pre-filtered student records
                        present_count = student_records_sem.filter(
                            Q(timetable__subject=cs) | Q(extra_class__subject=cs),
                            status__in=['Present', 'Late']
                        ).count()
                        absent_count = total_classes - present_count
                        context['subject_attendance_data'].append({
                            'subject_pk': cs.subject.pk, 'subject_name': cs.subject.name,
                            'attended_classes': present_count, 'absent_classes': absent_count,
                            'total_classes': total_classes,
                            'official_percentage': round(
                                (present_count / total_classes * 100) if total_classes > 0 else 0, 2)
                        })

                # Calculate overall semester stats after the loop
                context['overall_attended_sem'] = sum(d['attended_classes'] for d in context['subject_attendance_data'])
                total_sem = sum(d['total_classes'] for d in context['subject_attendance_data'])
                context['overall_absent_sem'] = total_sem - context['overall_attended_sem']
                context['overall_percentage_sem'] = round(
                    (context['overall_attended_sem'] / total_sem * 100) if total_sem > 0 else 0, 2)
                context['subject_attendance_data_json'] = json.dumps(context['subject_attendance_data'])

            elif view_type == 'monthly':
                # Get all records for the group in the semester to determine available months
                all_group_records = AttendanceRecord.objects.filter(
                    Q(timetable__student_group=student_group) | Q(extra_class__class_group=student_group),
                    Q(timetable__subject__semester=context['selected_semester']) | Q(
                        extra_class__subject__semester=context['selected_semester'])
                ).distinct()

                context['available_months'] = all_group_records.annotate(month=TruncMonth('date')).values(
                    'month').distinct().order_by('-month')

                if not selected_month_str and context['available_months']:
                    selected_month_str = context['available_months'][0]['month'].strftime('%Y-%m')
                context['selected_month'] = selected_month_str

                if selected_month_str:
                    year, month = map(int, selected_month_str.split('-'))
                    subjects_for_semester = all_course_subjects_qs.filter(semester=context['selected_semester'])
                    for cs in subjects_for_semester:
                        # Calculate total classes for the month (regular + extra)
                        regular_classes_month = AttendanceRecord.objects.filter(
                            timetable__student_group=student_group, timetable__subject=cs, date__year=year,
                            date__month=month
                        ).values('date', 'timetable__time_slot').distinct().count()

                        extra_classes_month = ExtraClass.objects.filter(
                            class_group=student_group, subject=cs, date__year=year, date__month=month
                        ).count()

                        total_classes_month = regular_classes_month + extra_classes_month

                        if total_classes_month > 0:
                            # Calculate present count for the month (regular + extra)
                            present_count_month = student_records_sem.filter(
                                Q(timetable__subject=cs) | Q(extra_class__subject=cs),
                                status__in=['Present', 'Late'], date__year=year, date__month=month
                            ).count()
                            absent_count_month = total_classes_month - present_count_month
                            context['monthly_subject_data'].append({
                                'subject_pk': cs.subject.pk, 'subject_name': cs.subject.name,
                                'attended_classes': present_count_month, 'absent_classes': absent_count_month,
                                'total_classes': total_classes_month,
                                'official_percentage': round(
                                    (present_count_month / total_classes_month * 100) if total_classes_month > 0 else 0,
                                    2)
                            })

                    # Calculate overall monthly stats after the loop
                    context['overall_attended_month'] = sum(
                        d['attended_classes'] for d in context['monthly_subject_data'])
                    total_month = sum(d['total_classes'] for d in context['monthly_subject_data'])
                    context['overall_absent_month'] = total_month - context['overall_attended_month']
                    context['overall_percentage_month'] = round(
                        (context['overall_attended_month'] / total_month * 100) if total_month > 0 else 0, 2)
                    context['monthly_subject_data_json'] = json.dumps(context['monthly_subject_data'])

            elif view_type == 'daily':
                try:
                    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
                    # Fetch all records for the day and prefetch related data for efficiency
                    daily_records = student_records_sem.filter(date=selected_date).select_related(
                        'timetable__subject__subject', 'timetable__time_slot',
                        'extra_class__subject__subject', 'extra_class__time_slot'
                    )

                    unified_daily_data = []
                    for record in daily_records:
                        if record.timetable:
                            unified_daily_data.append({
                                'subject_name': record.timetable.subject.subject.name,
                                'time_slot': record.timetable.time_slot,
                                'status': record.status,
                                'is_late': record.is_late
                            })
                        elif record.extra_class:
                            unified_daily_data.append({
                                'subject_name': record.extra_class.subject.subject.name,
                                'time_slot': record.extra_class.time_slot,
                                'status': record.status,
                                'is_late': record.is_late
                            })

                    # Sort the final list by the start time of the time slot
                    unified_daily_data.sort(key=lambda x: x['time_slot'].start_time)
                    context['daily_attendance_data'] = unified_daily_data

                except (ValueError, TypeError):
                    pass  # Ignore invalid date formats


            elif view_type == 'marks':

                # Pass the ordered queryset of marks directly to the template.

                # The .order_by() is important for grouping in the template.

                context['marks_data_list'] = Mark.objects.filter(

                    student=student,

                    subject__semester=context['selected_semester']

                ).select_related(

                    'subject__subject', 'criterion'

                ).order_by('subject__subject__name', 'criterion__name')

    return render(request, 'academics/admin_student_attendance_detail.html', context)


@login_required
@permission_required('academics.add_studentgroup', raise_exception=True)
def student_group_create_view(request):
    if request.method == 'POST':
        form = StudentGroupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'New class has been created successfully.')
            return redirect('academics:admin_select_class')
    else:
        form = StudentGroupForm()

    context = {
        'form': form,
        'form_title': 'Add New Class / Batch'
    }
    return render(request, 'academics/student_group_form.html', context)


@login_required
@permission_required('academics.update_studentgroup', raise_exception=True)
def student_group_update_view(request, pk):
    student_group = get_object_or_404(StudentGroup, pk=pk)
    if request.method == 'POST':
        form = StudentGroupForm(request.POST, instance=student_group)
        if form.is_valid():
            form.save()
            messages.success(request, f'Class "{student_group.name}" has been updated.')
            return redirect('academics:admin_select_class')
    else:
        form = StudentGroupForm(instance=student_group)

    context = {
        'form': form,
        'form_title': f'Edit Class: {student_group.name}'
    }
    return render(request, 'academics/student_group_form.html', context)


@login_required
@permission_required('academics.view_course', raise_exception=True)
@nav_item(title="Manage Courses", icon="iconsminds-books", url_name='academics:course_list',
          permission='academics.view_course', group='admin_management', order=10)
def course_list_view(request):
    courses = Course.objects.all()
    return render(request, 'academics/course_list.html', {'courses': courses})


@login_required
@permission_required('academics.add_course', raise_exception=True)
def course_create_view(request):
    # We define our inline formset here
    CourseSubjectFormSet = inlineformset_factory(
        Course, CourseSubject,
        fields=('subject', 'semester'),
        extra=1,  # Start with one extra form
        can_delete=True,
        widgets={
            'subject': forms.Select(attrs={'class': 'custom-select'}),
            'semester': forms.NumberInput(attrs={'class': 'form-control'}),
        }
    )

    if request.method == 'POST':
        form = CourseForm(request.POST)
        formset = CourseSubjectFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                course = form.save()
                formset.instance = course
                formset.save()
                messages.success(request, 'Course and subjects saved successfully.')
                return redirect('academics:course_list')
    else:
        form = CourseForm()
        formset = CourseSubjectFormSet()

    subjects_qs = Subject.objects.all().values('id', 'code', 'name')
    subjects_json = json.dumps(list(subjects_qs))

    context = {
        'form': form,
        'formset': formset,
        'subjects_json': subjects_json,
        'form_title': 'Create New Course'
    }
    return render(request, 'academics/course_form_with_subjects.html', context)


@login_required
@permission_required('academics.update_course', raise_exception=True)
def course_update_view(request, pk):
    course = get_object_or_404(Course, pk=pk)
    CourseSubjectFormSet = inlineformset_factory(
        Course, CourseSubject,
        fields=('subject', 'semester'),
        extra=1,
        can_delete=True,
        widgets={
            'subject': forms.Select(attrs={'class': 'form-control'}),
            'semester': forms.NumberInput(attrs={'class': 'form-control'}),
        }
    )

    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        formset = CourseSubjectFormSet(request.POST, instance=course)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                formset.save()
                messages.success(request, 'Course and subjects updated successfully.')
                return redirect('academics:course_list')
    else:
        form = CourseForm(instance=course)
        formset = CourseSubjectFormSet(instance=course)

    subjects_qs = Subject.objects.all().values('id', 'code', 'name')
    subjects_json = json.dumps(list(subjects_qs))

    context = {
        'form': form,
        'formset': formset,
        'subjects_json': subjects_json,
        'form_title': 'Create New Course'
    }
    return render(request, 'academics/course_form_with_subjects.html', context)


@login_required
@permission_required('academics.delete_course', raise_exception=True)
def course_delete_view(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        messages.success(request, 'Course deleted successfully.')
        return redirect('academics:course_list')
    return render(request, 'academics/course_confirm_delete.html', {'course': course})


@login_required
@permission_required('academics.delete_studentgroup', raise_exception=True)
def student_group_delete_view(request, pk):
    student_group = get_object_or_404(StudentGroup, pk=pk)
    if request.method == 'POST':
        student_group.delete()
        messages.success(request, 'Class deleted successfully.')
        return redirect('academics:admin_select_class')
    return render(request, 'academics/student_group_confirm_delete.html', {'student_group': student_group})


@login_required
@permission_required('academics.add_student', raise_exception=True)
@transaction.atomic
def student_create_view(request, group_id):
    student_group = get_object_or_404(StudentGroup, pk=group_id)
    if request.method == 'POST':
        form = AddStudentForm(request.POST, request.FILES)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data['username'], email=form.cleaned_data['email'],
                password=form.cleaned_data['password'], first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name']
            )
            profile = user.profile
            profile.role = 'student'
            profile.student_id_number = form.cleaned_data['student_id_number']
            profile.contact_number = form.cleaned_data['contact_number']
            profile.student_group = student_group
            profile.photo = form.cleaned_data.get('photo')
            profile.father_name = form.cleaned_data.get('father_name')
            profile.father_phone = form.cleaned_data.get('father_phone')
            profile.mother_name = form.cleaned_data.get('mother_name')
            profile.mother_phone = form.cleaned_data.get('mother_phone')
            profile.address = form.cleaned_data.get('address')
            profile.save()
            messages.success(request, f"Student '{user.username}' created.")
            if '_addanother' in request.POST:
                return redirect('academics:student_create', group_id=group_id)
            else:
                return redirect('academics:admin_student_list', group_id=group_id)
    else:
        form = AddStudentForm()
    context = {'form': form, 'form_title': f'Add New Student to {student_group.name}', 'student_group': student_group}
    return render(request, 'academics/student_form.html', context)


@login_required
@permission_required('academics.delete_student', raise_exception=True)
def student_delete_view(request, pk):
    student = get_object_or_404(User, pk=pk)
    # Find the group ID before deleting the student to redirect back correctly
    student_group = student.student_groups.first()
    group_id = student_group.id if student_group else None

    if request.method == 'POST':
        # The associated Profile will be deleted automatically due to on_delete=models.CASCADE
        student.delete()
        messages.success(request, 'Student deleted successfully.')
        if group_id:
            return redirect('academics:admin_student_list', group_id=group_id)
        return redirect('home')  # Fallback redirect

    return render(request, 'academics/student_confirm_delete.html', {'student': student})


@login_required
@permission_required('academics.delete_studentgroup', raise_exception=True)
def student_group_delete_view(request, pk):
    """
    Delete a student group (class).
    """
    student_group = get_object_or_404(StudentGroup, pk=pk)

    if request.method == 'POST':
        student_group.delete()
        messages.success(request, f'Class "{student_group}" has been deleted successfully.')
        return redirect('academics:course_list')

    return render(request, 'academics/confirm_delete.html', {
        'item': student_group,
        'type': 'Class',
    })


@login_required
@permission_required('academics.delete_student', raise_exception=True)
def student_delete_view(request, pk):
    student = get_object_or_404(User, pk=pk)
    student_group = student.profile.student_group  # To redirect back to the correct class list

    if request.method == 'POST':
        student.delete()
        messages.success(request, f'Student "{student.username}" has been deleted.')
        if student_group:
            return redirect('academics:admin_student_list', group_id=student_group.id)
        return redirect('academics:admin_select_class')  # Fallback

    return render(request, 'academics/confirm_delete.html', {'item': student, 'type': 'Student'})


@login_required
@permission_required('academics.view_subject', raise_exception=True)
@nav_item(title="Manage Subjects", icon="iconsminds-book", url_name='academics:subject_list',
          permission='academics.view_subject', group='admin_management', order=20)
def subject_list_view(request):
    subjects = Subject.objects.all().order_by('name')
    return render(request, 'academics/subject_list.html', {'subjects': subjects})


@login_required
@permission_required('academics.add_subject', raise_exception=True)
def subject_create_view(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject created successfully.')
            return redirect('academics:subject_list')
    else:
        form = SubjectForm()
    return render(request, 'academics/subject_form.html', {'form': form, 'form_title': 'Create New Subject'})


@login_required
@permission_required('academics.update_subject', raise_exception=True)
def subject_update_view(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully.')
            return redirect('academics:subject_list')
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'academics/subject_form.html', {'form': form, 'form_title': f'Edit Subject: {subject.name}'})


@login_required
@permission_required('academics.delete_subject', raise_exception=True)
def subject_delete_view(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, 'Subject deleted successfully.')
        return redirect('academics:subject_list')
    return render(request, 'academics/confirm_delete.html', {'item': subject, 'type': 'Subject'})


@login_required
@permission_required('academics.update_student', raise_exception=True)
def student_update_view(request, pk):
    student_user = get_object_or_404(User, pk=pk, profile__role='student')
    if request.method == 'POST':
        form = EditStudentForm(request.POST, request.FILES, instance=student_user.profile)
        if form.is_valid():
            form.save()
            student_user.first_name = form.cleaned_data['first_name']
            student_user.last_name = form.cleaned_data['last_name']
            student_user.email = form.cleaned_data['email']
            student_user.save()
            messages.success(request, 'Student details updated successfully.')
            return redirect('academics:admin_student_list', group_id=student_user.profile.student_group.id)
    else:
        form = EditStudentForm(instance=student_user.profile, initial={
            'first_name': student_user.first_name, 'last_name': student_user.last_name, 'email': student_user.email
        })
    context = {'form': form, 'form_title': f'Edit Student: {student_user.get_full_name()}', 'student': student_user}
    return render(request, 'academics/student_form.html', context)


@login_required
@permission_required('academics.delete_timeslot', raise_exception=True)
def timeslot_delete_view(request, pk):
    timeslot = get_object_or_404(TimeSlot, pk=pk)
    if request.method == 'POST':
        timeslot.delete()
        messages.success(request, "Time slot has been deleted.")
    return redirect('academics:admin_settings')


@login_required
@permission_required('academics.add_attendancerecord', raise_exception=True)
@nav_item(title="Daily Schedule", icon="simple-icon-check", url_name="academics:faculty_schedule",
          permission='academics.add_attendancerecord', group='faculty_tools', order=10)
def faculty_daily_schedule_view(request):
    today = timezone.now().date()
    current_day = today.strftime('%A')

    # Get attendance settings for deadline calculation
    attendance_settings = AttendanceSettings.load()
    mark_deadline = attendance_settings.mark_deadline_days

    # Get regular timetable entries for today
    regular_schedule = Timetable.objects.filter(
        faculty=request.user,
        day_of_week=current_day
    ).select_related('subject__subject', 'student_group', 'time_slot')

    # Get extra classes for today
    extra_classes = ExtraClass.objects.filter(
        teacher=request.user,
        date=today
    ).select_related('subject__subject', 'class_group', 'time_slot')

    schedule = []

    # Process regular schedule
    for entry in regular_schedule:
        # Check if there's a substitution for today
        substitution = DailySubstitution.objects.filter(timetable=entry, date=today).first()

        # Skip if someone else is substituting for this class today
        if substitution and substitution.substituted_by != request.user:
            continue

        # Check if class is manually cancelled for today
        is_cancelled = ClassCancellation.objects.filter(timetable=entry, date=today).exists()

        # Check if attendance has been marked for TODAY
        attendance_marked = AttendanceRecord.objects.filter(
            timetable=entry,
            date=today
        ).exists()

        schedule.append({
            'session': entry,
            'type': 'regular',
            'is_substitution': substitution is not None,
            'is_cancelled': is_cancelled,
            'attendance_marked': attendance_marked,
        })

    # Process extra classes (similar logic but simpler since they're one-time events)
    for extra_class in extra_classes:
        attendance_marked = AttendanceRecord.objects.filter(
            extra_class=extra_class,
            date=today
        ).exists()

        schedule.append({
            'session': extra_class,
            'type': 'extra',
            'is_substitution': False,
            'is_cancelled': False,  # Extra classes are typically not auto-cancelled
            'attendance_marked': attendance_marked,
        })

    # Sort by time slot
    schedule.sort(key=lambda x: x['session'].time_slot.start_time)

    context = {
        'schedule': schedule,
        'current_day': current_day,
        'mark_deadline_days': mark_deadline,
    }
    return render(request, 'academics/faculty_schedule.html', context)


@login_required
@permission_required('academics.add_attendancerecord', raise_exception=True)
def mark_extra_class_attendance_view(request, extra_class_id):
    current_date = timezone.now().date()
    extra_class_entry = get_object_or_404(ExtraClass, pk=extra_class_id)
    print(extra_class_entry)

    # Security check: ensure the logged-in user is the teacher for this extra class
    if extra_class_entry.teacher != request.user:
        raise PermissionDenied("You are not authorized to mark attendance for this class.")

    students = User.objects.filter(profile__student_group=extra_class_entry.class_group).order_by('first_name')
    MarkAttendanceFormSet = formset_factory(MarkAttendanceForm, extra=0)

    if request.method == 'POST':
        formset = MarkAttendanceFormSet(request.POST)
        if formset.is_valid():
            with transaction.atomic():
                for form in formset:
                    student_id = form.cleaned_data['student_id']
                    status = form.cleaned_data['status']
                    is_late = form.cleaned_data.get('is_late', False)
                    student = User.objects.get(pk=student_id)

                    # --- CHANGE: Save record with extra_class foreign key ---
                    record, created = AttendanceRecord.objects.update_or_create(
                        student=student,
                        extra_class=extra_class_entry,
                        date=current_date,
                        defaults={'status': status, 'is_late': is_late, 'marked_by': request.user}
                    )
            messages.success(request, "Attendance for the extra class has been marked successfully.")
            return redirect('academics:faculty_schedule')
    else:
        initial_data = []
        existing_records = AttendanceRecord.objects.filter(extra_class=extra_class_entry, date=current_date)
        status_map = {record.student_id: (record.status, record.is_late) for record in existing_records}
        for student in students:
            saved_status, saved_is_late = status_map.get(student.id, ('Present', False))
            initial_data.append({'student_id': student.id, 'status': saved_status, 'is_late': saved_is_late})
        formset = MarkAttendanceFormSet(initial=initial_data)

    student_forms = zip(students, formset)
    context = {
        # --- FIX: Use 'timetable_entry' as the key to match the reused template ---
        'session_object': extra_class_entry,
        'formset': formset,
        'student_forms': student_forms,
        'date': current_date,
        'is_extra_class': True  # This flag is still useful for conditional logic in the template
    }
    # We can reuse the existing mark_attendance.html template
    return render(request, 'academics/mark_attendance.html', context)


@login_required
@permission_required('academics.add_attendancerecord', raise_exception=True)
def mark_attendance_view(request, timetable_id, date_str=None):
    if date_str:
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        current_date = timezone.now().date()

    timetable_entry = get_object_or_404(Timetable, pk=timetable_id)
    students = User.objects.filter(profile__student_group=timetable_entry.student_group).order_by('first_name')

    MarkAttendanceFormSet = formset_factory(MarkAttendanceForm, extra=0)

    is_cancelled = ClassCancellation.objects.filter(timetable=timetable_entry, date=current_date).exists()
    if is_cancelled:
        messages.warning(request, "This class session was marked as not conducted, so attendance cannot be taken.")
        return redirect('academics:faculty_schedule')

    if request.method == 'POST':
        formset = MarkAttendanceFormSet(request.POST)
        if formset.is_valid():
            with transaction.atomic():
                for form in formset:
                    student_id = form.cleaned_data['student_id']
                    status = form.cleaned_data['status']
                    is_late = form.cleaned_data.get('is_late', False)  # Get the checkbox value
                    student = User.objects.get(pk=student_id)

                    record, created = AttendanceRecord.objects.get_or_create(
                        student=student,
                        timetable=timetable_entry,
                        date=current_date,
                        defaults={'status': status, 'is_late': is_late, 'marked_by': request.user}
                    )

                    if not created:
                        record.status = status
                        record.is_late = is_late  # Update the is_late field on edit
                        record.marked_by = request.user
                        record.save()

            messages.success(request, "Attendance has been marked successfully.")
            return redirect('academics:faculty_schedule')
    else:
        # This logic now also fetches the is_late status to pre-fill the form
        initial_data = []
        existing_records = AttendanceRecord.objects.filter(timetable=timetable_entry, date=current_date)
        status_map = {record.student_id: (record.status, record.is_late) for record in existing_records}
        for student in students:
            saved_status, saved_is_late = status_map.get(student.id, ('Present', False))
            initial_data.append({
                'student_id': student.id,
                'status': saved_status,
                'is_late': saved_is_late,
            })
        formset = MarkAttendanceFormSet(initial=initial_data)

    student_forms = zip(students, formset)
    context = {
        'session_object': timetable_entry,
        'formset': formset,
        'student_forms': student_forms,
        'date': current_date,
        'is_extra_class': False,
    }
    return render(request, 'academics/mark_attendance.html', context)


@login_required
@permission_required('academics.add_classcancellation', raise_exception=True)
def class_cancellation_view(request, timetable_id):
    timetable_entry = get_object_or_404(Timetable, pk=timetable_id)
    if request.method == 'POST':
        # Create a cancellation record for this class on this day
        ClassCancellation.objects.get_or_create(
            timetable=timetable_entry,
            date=timezone.now().date(),
            defaults={'cancelled_by': request.user}
        )
        messages.info(request,
                      f"Class '{timetable_entry.subject.subject.name}' has been marked as not conducted for today.")
        return redirect('academics:faculty_schedule')

    # Redirect if accessed via GET
    return redirect('academics:faculty_schedule')


@login_required
@permission_required('academics.view_attendancerecord', raise_exception=True)
@nav_item(title="Daily Activity Log", icon="iconsminds-folder-open", url_name="academics:daily_log",
          permission='academics.view_attendancerecord', group='admin_management', order=40)
def daily_attendance_log_view(request):
    # Get all distinct dates from regular attendance, extra class attendance, and cancellations
    attendance_dates = AttendanceRecord.objects.values_list('date', flat=True)
    extra_class_dates = ExtraClass.objects.values_list('date', flat=True)
    cancellation_dates = ClassCancellation.objects.values_list('date', flat=True)
    all_dates = sorted(list(set(chain(attendance_dates, extra_class_dates, cancellation_dates))), reverse=True)

    log_data = {}
    for date in all_dates:
        # --- MODIFIED LOGIC TO UNIFY SESSIONS ---
        unified_daily_log = []

        # Get substitutions for the date
        subs = DailySubstitution.objects.filter(date=date).select_related('substituted_by')
        sub_map = {s.timetable_id: s.substituted_by for s in subs}

        # 1. Get conducted regular classes
        conducted_regular_entries = Timetable.objects.filter(attendance_records__date=date).distinct()
        for entry in conducted_regular_entries:
            present_count = AttendanceRecord.objects.filter(timetable=entry, date=date, status='Present').count()
            total_students = Profile.objects.filter(student_group=entry.student_group).count()
            unified_daily_log.append({
                'session': entry, 'type': 'regular', 'status': 'Conducted',
                'present_count': present_count, 'total_students': total_students,
                'date': date, 'substituted_by': sub_map.get(entry.id)
            })

        # 2. Get conducted extra classes
        conducted_extra_entries = ExtraClass.objects.filter(attendance_records__date=date).distinct()
        for entry in conducted_extra_entries:
            present_count = AttendanceRecord.objects.filter(extra_class=entry, date=date, status='Present').count()
            total_students = Profile.objects.filter(student_group=entry.class_group).count()
            unified_daily_log.append({
                'session': entry, 'type': 'extra', 'status': 'Conducted',
                'present_count': present_count, 'total_students': total_students, 'date': date
            })

        # 3. Get cancelled regular classes
        cancelled_entries = Timetable.objects.filter(cancellations__date=date).distinct()
        for entry in cancelled_entries:
            # Avoid showing a class as both conducted and cancelled if data is inconsistent
            if not any(log['session'] == entry and log['type'] == 'regular' for log in unified_daily_log):
                unified_daily_log.append({'session': entry, 'type': 'regular', 'status': 'Cancelled', 'date': date})

        # Sort all log entries for the day by time
        unified_daily_log.sort(key=lambda x: x['session'].time_slot.start_time)
        log_data[date] = unified_daily_log

    context = {'log_data': log_data}
    return render(request, 'academics/daily_attendance_log.html', context)


@login_required
@permission_required('academics.view_attendancelog', raise_exception=True)
def daily_log_detail_view(request, timetable_id, date):
    """
    Shows the detailed attendance list for a single REGULAR class session.
    """
    try:
        timetable_entry = Timetable.objects.select_related(
            'subject__subject', 'student_group', 'faculty'
        ).get(pk=timetable_id)

        records = AttendanceRecord.objects.filter(
            timetable=timetable_entry,
            date=date
        ).select_related('student').order_by('student__first_name')

    except ObjectDoesNotExist:
        raise Http404("The requested attendance session does not exist.")

    context = {
        # MODIFICATION: Use a generic context variable name 'session_entry'
        'session_entry': timetable_entry,
        'date': date,
        'records': records,
        'is_extra_class': False,
    }
    return render(request, 'academics/daily_log_detail.html', context)


# --- ADD THIS ENTIRE NEW VIEW FUNCTION ---
@login_required
@permission_required('academics.view_attendancerecord')
def extra_class_log_detail_view(request, extra_class_id, date):
    """
    Shows the detailed attendance list for a single EXTRA class session.
    """
    try:
        # Fetch the specific extra class session
        extra_class_entry = ExtraClass.objects.select_related(
            'subject__subject', 'class_group', 'teacher'
        ).get(pk=extra_class_id)

        # Fetch the attendance records for that session on that date
        records = AttendanceRecord.objects.filter(
            extra_class=extra_class_entry,
            date=date
        ).select_related('student').order_by('student__first_name')

    except ObjectDoesNotExist:
        raise Http404("The requested extra class session does not exist.")

    context = {
        # Use the same generic context variable name for template reuse
        'session_entry': extra_class_entry,
        'date': date,
        'records': records,
        'is_extra_class': True,
    }
    # Reuse the same template as the regular detail view
    return render(request, 'academics/daily_log_detail.html', context)


@login_required
@permission_required('academics.add_timetable', raise_exception=True)
def timetable_entry_create_view(request, group_id, day, timeslot_id):
    student_group = get_object_or_404(StudentGroup, pk=group_id)
    time_slot = get_object_or_404(TimeSlot, pk=timeslot_id)  # Fixed variable name

    if request.method == 'POST':
        form = TimetableEntryForm(request.POST, student_group=student_group)
        if form.is_valid():
            try:
                entry = form.save(commit=False)
                entry.student_group = student_group
                entry.day_of_week = day
                entry.time_slot = time_slot
                entry.save()
                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = TimetableEntryForm(student_group=student_group)

    context = {
        'form': form,
        'form_title': f'Add Class for {student_group.name} - {day} at {time_slot}',
        'student_group': student_group,
        'day': day,
        'time_slot': time_slot,
    }

    # Always return the modal form for timetable operations
    return render(request, 'partials/modal_form.html', context)


@login_required
@permission_required('academics.change_timetable', raise_exception=True)
def timetable_entry_update_view(request, entry_id):
    entry = get_object_or_404(Timetable, pk=entry_id)

    if request.method == 'POST':
        form = TimetableEntryForm(request.POST, instance=entry, student_group=entry.student_group)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = TimetableEntryForm(instance=entry, student_group=entry.student_group)

    context = {
        'form': form,
        'form_title': f'Edit Class - {entry.subject.subject.name}',
        'entry': entry,
    }

    return render(request, 'partials/modal_form.html', context)


@login_required
@permission_required('academics.delete_timetable', raise_exception=True)
def timetable_entry_delete_view(request, entry_id):
    entry = get_object_or_404(Timetable, pk=entry_id)

    if request.method == 'POST':
        try:
            entry.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@permission_required('academics.view_timetable', raise_exception=True)
@nav_item(title="Manage Timetable", icon="iconsminds-calendar-4", url_name="academics:manage_timetable",
          permission='academics.view_timetable', group='admin_management', order=40)
def manage_timetable_view(request):
    student_groups = StudentGroup.objects.all()
    selected_group = None
    timetable_grid = {}

    group_id = request.GET.get('group_id')
    if group_id:
        selected_group = get_object_or_404(StudentGroup, pk=group_id)

        entries = Timetable.objects.filter(student_group=selected_group).select_related('subject__subject', 'faculty',
                                                                                        'time_slot')

        for entry in entries:
            if entry.day_of_week not in timetable_grid:
                timetable_grid[entry.day_of_week] = {}
            timetable_grid[entry.day_of_week][entry.time_slot.id] = entry

    context = {
        'student_groups': student_groups,
        'selected_group': selected_group,
        'timeslots': TimeSlot.objects.all(),
        'days_of_week': [day[0] for day in Timetable.DAY_CHOICES],
        'timetable_grid': timetable_grid,
    }
    return render(request, 'academics/manage_timetable.html', context)


@login_required
@permission_required('academics.add_dailysubstitution', raise_exception=True)
@nav_item(title="Manage Substitutions", icon="iconsminds-shuffle-1", url_name="academics:manage_substitutions",
          group='admin_management', permission='academics.add_dailysubstitution', order=50)
def manage_substitutions_view(request):
    # Get the date from the request's GET parameters, default to today if not provided
    selected_date_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    day_of_week = selected_date.strftime('%A')

    # Get the schedule and existing substitutions for the selected date
    todays_schedule = Timetable.objects.filter(day_of_week=day_of_week).order_by('time_slot')
    existing_substitutions = DailySubstitution.objects.filter(date=selected_date).select_related('substituted_by')

    # Create a dictionary for easy lookup in the template
    substitutions_map = {sub.timetable_id: sub.substituted_by for sub in existing_substitutions}

    context = {
        'selected_date': selected_date,
        'schedule': todays_schedule,
        'substitutions_map': substitutions_map,
        'substitution_form': SubstitutionForm()  # An instance of our new form
    }
    return render(request, 'academics/manage_substitutions.html', context)


@login_required
@permission_required('academics.add_dailysubstitution', raise_exception=True)
def assign_substitution_view(request, timetable_id):
    if request.method == 'POST':
        form = SubstitutionForm(request.POST)
        if form.is_valid():
            timetable_entry = get_object_or_404(Timetable, pk=timetable_id)
            substitute_faculty = form.cleaned_data['substitute_faculty']
            date = request.POST.get('date')  # Get the date from the POST data

            # Use update_or_create to handle both new and existing substitutions
            DailySubstitution.objects.update_or_create(
                timetable=timetable_entry,
                date=date,
                defaults={'substituted_by': substitute_faculty}
            )
            messages.success(request, f"Substitution saved for {timetable_entry.subject.subject.name}.")
        else:
            messages.error(request, "Invalid selection. Please try again.")

    # Redirect back to the substitution page for the same date
    return redirect(f"{reverse('academics:manage_substitutions')}?date={request.POST.get('date', '')}")


@login_required
@permission_required('academics.delete_dailysubstitution', raise_exception=True)
def cancel_substitution_view(request, timetable_id):
    if request.method == 'POST':
        date_str = request.POST.get('date')
        # Find the specific substitution and delete it
        substitution = get_object_or_404(DailySubstitution, timetable_id=timetable_id, date=date_str)
        substitution.delete()
        messages.success(request, "Substitution has been cancelled successfully.")
        # Redirect back to the same date on the management page
        return redirect(f"{reverse('academics:manage_substitutions')}?date={date_str}")

    # Redirect to the main page if accessed via GET
    return redirect('academics:manage_substitutions')


@login_required
@permission_required('academics.view_own_attendance', raise_exception=True)
@nav_item(title="My Attendance", icon="simple-icon-pie-chart", url_name="academics:student_my_attendance",
          permission='academics.view_own_attendance', group='my_academics', order=1, role_required='student')
def student_my_attendance_view(request):
    """
    Allows a logged-in student to view their own attendance details,
    filtered by semester, with charts for visualization. This view now
    correctly includes attendance from both regular and extra classes.
    """
    student = request.user
    student_group = student.profile.student_group if hasattr(student, 'profile') else None

    # --- Get view type and filters from request ---
    view_type = request.GET.get('view_type', 'semester')
    selected_semester = request.GET.get('semester')
    selected_month_str = request.GET.get('month')
    selected_date_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))

    # --- Initialize a comprehensive context dictionary ---
    context = {
        'student': student,
        'view_type': view_type,
        'available_semesters': [],
        'selected_semester': None,
        'subject_attendance_data': [],
        'subject_attendance_data_json': '[]',
        'overall_attended_sem': 0,
        'overall_absent_sem': 0,
        'overall_percentage_sem': 0,
        'monthly_subject_data': [],
        'monthly_subject_data_json': '[]',
        'available_months': [],
        'selected_month': selected_month_str,
        'overall_attended_month': 0,
        'overall_absent_month': 0,
        'overall_percentage_month': 0,
        'daily_attendance_data': [],
        'selected_date': selected_date_str,
        'marks_data_list': []
    }

    if student_group and student_group.course:
        all_course_subjects_qs = CourseSubject.objects.filter(course=student_group.course)
        context['available_semesters'] = all_course_subjects_qs.values_list('semester', flat=True).distinct().order_by(
            'semester')

        if not selected_semester and context['available_semesters']:
            selected_semester = context['available_semesters'].last()
        context['selected_semester'] = int(selected_semester) if selected_semester and str(
            selected_semester).isdigit() else None

        if context['selected_semester']:
            # Base queryset for ALL of the student's attendance records for the semester
            student_records_sem = AttendanceRecord.objects.filter(
                student=student
            ).filter(
                Q(timetable__subject__semester=context['selected_semester']) |
                Q(extra_class__subject__semester=context['selected_semester'])
            )

            if view_type == 'semester':
                subjects_for_semester = all_course_subjects_qs.filter(semester=context['selected_semester'])
                for cs in subjects_for_semester:
                    # Calculate total classes by combining regular and extra classes
                    regular_classes_count = AttendanceRecord.objects.filter(
                        timetable__student_group=student_group, timetable__subject=cs
                    ).values('date', 'timetable__time_slot').distinct().count()
                    extra_classes_count = ExtraClass.objects.filter(
                        class_group=student_group, subject=cs
                    ).count()
                    total_classes = regular_classes_count + extra_classes_count

                    if total_classes > 0:
                        # Count attended classes from both regular and extra class records
                        present_count = student_records_sem.filter(
                            Q(timetable__subject=cs) | Q(extra_class__subject=cs),
                            status__in=['Present', 'Late']
                        ).count()
                        absent_count = total_classes - present_count
                        context['subject_attendance_data'].append({
                            'subject_pk': cs.subject.pk,
                            'subject_name': cs.subject.name,
                            'attended_classes': present_count,
                            'absent_classes': absent_count,
                            'total_classes': total_classes,
                            'official_percentage': round(
                                (present_count / total_classes * 100) if total_classes > 0 else 0, 2)
                        })

                context['overall_attended_sem'] = sum(d['attended_classes'] for d in context['subject_attendance_data'])
                total_sem = sum(d['total_classes'] for d in context['subject_attendance_data'])
                context['overall_absent_sem'] = total_sem - context['overall_attended_sem']
                context['overall_percentage_sem'] = round(
                    (context['overall_attended_sem'] / total_sem * 100) if total_sem > 0 else 0, 2)
                context['subject_attendance_data_json'] = json.dumps(context['subject_attendance_data'])

            elif view_type == 'monthly':
                all_records_in_sem = AttendanceRecord.objects.filter(
                    Q(timetable__student_group=student_group) | Q(extra_class__class_group=student_group),
                    Q(timetable__subject__semester=context['selected_semester']) | Q(
                        extra_class__subject__semester=context['selected_semester'])
                ).distinct()
                context['available_months'] = all_records_in_sem.annotate(month=TruncMonth('date')).values(
                    'month').distinct().order_by('-month')

                if not selected_month_str and context['available_months']:
                    selected_month_str = context['available_months'][0]['month'].strftime('%Y-%m')
                context['selected_month'] = selected_month_str

                if selected_month_str:
                    year, month = map(int, selected_month_str.split('-'))
                    subjects_for_semester = all_course_subjects_qs.filter(semester=context['selected_semester'])
                    for cs in subjects_for_semester:
                        regular_classes_month = AttendanceRecord.objects.filter(
                            timetable__student_group=student_group, timetable__subject=cs, date__year=year,
                            date__month=month
                        ).values('date', 'timetable__time_slot').distinct().count()
                        extra_classes_month = ExtraClass.objects.filter(
                            class_group=student_group, subject=cs, date__year=year, date__month=month
                        ).count()
                        total_classes_month = regular_classes_month + extra_classes_month

                        if total_classes_month > 0:
                            present_count_month = student_records_sem.filter(
                                Q(timetable__subject=cs) | Q(extra_class__subject=cs),
                                status__in=['Present', 'Late'], date__year=year, date__month=month
                            ).count()
                            absent_count_month = total_classes_month - present_count_month
                            context['monthly_subject_data'].append({
                                'subject_pk': cs.subject.pk,
                                'subject_name': cs.subject.name,
                                'attended_classes': present_count_month,
                                'absent_classes': absent_count_month,
                                'total_classes': total_classes_month,
                                'official_percentage': round(
                                    (present_count_month / total_classes_month * 100) if total_classes_month > 0 else 0,
                                    2)
                            })

                    context['overall_attended_month'] = sum(
                        d['attended_classes'] for d in context['monthly_subject_data'])
                    total_month = sum(d['total_classes'] for d in context['monthly_subject_data'])
                    context['overall_absent_month'] = total_month - context['overall_attended_month']
                    context['overall_percentage_month'] = round(
                        (context['overall_attended_month'] / total_month * 100) if total_month > 0 else 0, 2)
                    context['monthly_subject_data_json'] = json.dumps(context['monthly_subject_data'])

            elif view_type == 'daily':
                try:
                    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
                    daily_records = student_records_sem.filter(date=selected_date).select_related(
                        'timetable__subject__subject', 'timetable__time_slot',
                        'extra_class__subject__subject', 'extra_class__time_slot'
                    )

                    unified_daily_data = []
                    for record in daily_records:
                        session_info = {
                            'status': record.status,
                            'is_late': record.is_late
                        }
                        if record.timetable:
                            session_info['subject_name'] = record.timetable.subject.subject.name
                            session_info['time_slot'] = record.timetable.time_slot
                        elif record.extra_class:
                            session_info['subject_name'] = record.extra_class.subject.subject.name
                            session_info['time_slot'] = record.extra_class.time_slot

                        if 'time_slot' in session_info:
                            unified_daily_data.append(session_info)

                    unified_daily_data.sort(key=lambda x: x['time_slot'].start_time)
                    context['daily_attendance_data'] = unified_daily_data
                except (ValueError, TypeError):
                    pass  # Ignore invalid date formats

            elif view_type == 'marks':
                context['marks_data_list'] = Mark.objects.filter(
                    student=student,
                    subject__semester=context['selected_semester']
                ).select_related(
                    'subject__subject', 'criterion'
                ).order_by('subject__subject__name', 'criterion__name')

    return render(request, 'academics/student_my_attendance.html', context)


@login_required
@permission_required('academics.change_attendancerecord', raise_exception=True)
@nav_item(title="Edit Past Attendance", icon="simple-icon-note", url_name="academics:previous_attendance",
          permission='academics.change_attendancerecord', group='faculty_tools', order=20)
def previous_attendance_view(request):
    settings = AttendanceSettings.load()
    # Calculate the cutoff date based on the edit deadline
    cutoff_date = timezone.now().date() - timezone.timedelta(days=settings.edit_deadline_days)

    # Find all distinct sessions marked by this faculty within the deadline
    marked_sessions = AttendanceRecord.objects.filter(
        marked_by=request.user,
        date__gte=cutoff_date
    ).values('date', 'timetable_id').distinct().order_by('-date')

    # Fetch the related timetable details for display
    session_details = []
    for session in marked_sessions:
        entry = Timetable.objects.select_related('time_slot', 'subject__subject', 'student_group').get(
            pk=session['timetable_id'])
        session_details.append({
            'date': session['date'],
            'timetable': entry
        })

    context = {'sessions': session_details}
    return render(request, 'academics/previous_attendance.html', context)


@login_required
@permission_required('academics.view_own_timetable', raise_exception=True)  # Re-using existing student permission
@nav_item(title="My Timetable", icon="iconsminds-calendar-4", url_name="academics:student_timetable",
          permission='academics.view_own_timetable', group='my_academics', order=2)
def student_timetable_view(request):
    """
    Displays the weekly timetable and any extra classes for the current day
    for the logged-in student's class.
    """
    # Safely get the student's assigned group from their profile.
    student_group = None
    try:
        student_group = request.user.profile.student_group
    except Profile.DoesNotExist:
        messages.error(request,
                       "CRITICAL: A profile for your user account does not exist. Please contact an administrator.")
        # Return a completely empty page if there's no profile
        return render(request, 'academics/student_timetable.html', {'timetable_grid': {}})

    timetable_grid = {}
    extra_classes_today = []

    # Only try to fetch timetable and extra classes if the student is assigned to a group.
    if student_group:
        # Get the regular weekly timetable entries
        entries = Timetable.objects.filter(student_group=student_group).select_related(
            'subject__subject', 'faculty', 'time_slot'
        )
        for entry in entries:
            if entry.day_of_week not in timetable_grid:
                timetable_grid[entry.day_of_week] = {}
            timetable_grid[entry.day_of_week][entry.time_slot.id] = entry

        # Get extra classes for today
        current_date = timezone.now().date()
        extra_classes_today = ExtraClass.objects.filter(
            class_group=student_group,
            date=current_date
        ).select_related('time_slot', 'subject__subject', 'teacher').order_by('time_slot__start_time')
    else:
        # If the student is assigned a profile but no group, show a clear warning.
        messages.warning(request,
                         "You are not currently assigned to any class group. Please contact the admin to see your full timetable.")

    context = {
        'student_group': student_group,
        'timeslots': TimeSlot.objects.filter(is_schedulable=True),
        'days_of_week': [day[0] for day in Timetable.DAY_CHOICES],
        'timetable_grid': timetable_grid,
        'extra_classes_today': extra_classes_today,  # Add extra classes to the context
    }
    return render(request, 'academics/student_timetable.html', context)


@login_required
@permission_required('academics.view_attendancerecord')  # Admin permission
@nav_item(title="Attendance Reports", icon="simple-icon-printer", url_name="academics:attendance_report",
          permission='academics.view_attendancerecord', group='admin_management')
def attendance_report_view(request):
    form = AttendanceReportForm()
    return render(request, 'academics/attendance_report_form.html', {'form': form})


@login_required
@permission_required('academics.view_attendancerecord')
def download_attendance_report_view(request):
    group_id = request.GET.get('student_group')
    subject_id = request.GET.get('subject')
    month = int(request.GET.get('month'))
    year = int(request.GET.get('year'))

    student_group = StudentGroup.objects.get(pk=group_id)
    subject = Subject.objects.get(pk=subject_id)
    students = User.objects.filter(profile__student_group=student_group).order_by('last_name', 'first_name')

    # Fetch all attendance records for the specific subject, class, and month
    records = AttendanceRecord.objects.filter(
        timetable__student_group=student_group,
        timetable__subject__subject=subject,
        date__year=year,
        date__month=month
    ).order_by('date')  # Order by date to process chronologically

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{subject.name} - {student_group.name}"

    # Create Header Row for ALL days in the month
    num_days = calendar.monthrange(year, month)[1]
    headers = ["Student Name"] + list(range(1, num_days + 1))
    sheet.append(headers)
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # --- FIX: Manually ensure uniqueness in Python instead of the DB ---
    records_dict = {}
    processed_keys = set()  # Keep track of (student_id, day) pairs we've already handled
    for record in records:
        key = (record.student.id, record.date.day)
        if key not in processed_keys:
            records_dict[key] = record.status[0]
            processed_keys.add(key)
    # --- END FIX ---

    # Populate Data Rows
    for student in students:
        row_data = [student.get_full_name()]
        for day in range(1, num_days + 1):
            status = records_dict.get((student.id, day), "-")
            row_data.append(status)
        sheet.append(row_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename="Attendance_{subject.name}_{student_group.name}_{month}_{year}.xlsx"'
    workbook.save(response)

    return response


@login_required
@permission_required('academics.can_send_announcement')
@nav_item(title="Announcements", icon="simple-icon-speech", url_name="academics:announcement_list",
          permission='academics.can_send_announcement', group='admin_management')
def announcement_list_view(request):
    announcements = Announcement.objects.all()
    return render(request, 'academics/announcement_list.html', {'announcements': announcements})


@login_required
@permission_required('academics.can_send_announcement')
@transaction.atomic
def announcement_create_view(request):
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            # Create the object in memory but don't save to DB yet
            announcement = form.save(commit=False)
            announcement.sender = request.user

            # Now save the main object and its M2M relationships.
            # The transaction ensures this is an all-or-nothing operation.
            announcement.save()
            form.save_m2m()

            messages.success(request, "Announcement has been sent successfully.")
            return redirect('academics:announcement_list')
    else:
        form = AnnouncementForm()
    return render(request, 'academics/announcement_form.html', {'form': form})


@login_required
def check_announcements_view(request):
    """
    Checks for the single latest unread announcement for the user.
    If found, returns its data and marks it as seen to prevent future pop-ups.
    """
    user = request.user

    # Admins don't get pop-ups
    if not hasattr(user, 'profile') or user.profile.role == 'admin':
        return JsonResponse({'announcement': None})

    # Determine which announcements are relevant to this user
    target_filter = Q()
    if user.profile.role == 'student' and user.profile.student_group:
        target_filter = Q(send_to_all_students=True) | Q(target_student_groups=user.profile.student_group)
    elif user.profile.role == 'faculty':
        target_filter = Q(send_to_all_faculty=True)

    if not target_filter:
        return JsonResponse({'announcement': None})

    # Find the latest announcement that the user has NOT seen yet
    try:
        latest_unread = Announcement.objects.filter(target_filter).exclude(
            usernotificationstatus__user=user
        ).latest('created_at')  # .latest() gets the single newest one

        # Mark this announcement as seen immediately so it doesn't pop up again
        UserNotificationStatus.objects.create(user=user, announcement=latest_unread)

        # Prepare the data to send to the frontend
        announcement_data = {
            'title': latest_unread.title,
            'content': latest_unread.content,
            'timestamp': latest_unread.created_at.strftime('%b %d, %Y, %I:%M %p'),
        }
        return JsonResponse({'announcement': announcement_data})

    except Announcement.DoesNotExist:
        # No unread announcements found, which is a normal case
        return JsonResponse({'announcement': None})


@login_required
def global_search_view(request):
    """
    Handles the global search query from the top navigation bar.
    Searches for Students, Faculty, and Subjects.
    """
    query = request.GET.get('q', '')
    students_found = User.objects.none()
    faculty_found = User.objects.none()
    subjects_found = Subject.objects.none()

    if query:
        # Search for students by name, username, or student ID
        students_found = User.objects.filter(
            Q(profile__role='student') &
            (Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(username__icontains=query) | Q(
                profile__student_id_number__icontains=query))
        ).distinct()

        # Search for faculty by name or username
        faculty_found = User.objects.filter(
            Q(profile__role='faculty') &
            (Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(username__icontains=query))
        ).distinct()

        # Search for subjects by name or code
        subjects_found = Subject.objects.filter(
            Q(name__icontains=query) | Q(code__icontains=query)
        )

    context = {
        'query': query,
        'students': students_found,
        'faculty': faculty_found,
        'subjects': subjects_found,
        'result_count': len(students_found) + len(faculty_found) + len(subjects_found)
    }
    return render(request, 'academics/search_results.html', context)


@login_required
@permission_required('academics.view_attendancerecord', raise_exception=True)
@nav_item(title="Late Comers", icon="simple-icon-clock", url_name="academics:late_comers",
          permission='academics.view_attendancerecord', group='admin_management', order=60)
def late_comers_view(request):
    student_groups = StudentGroup.objects.all()
    selected_group_id = request.GET.get('student_group')
    selected_month = request.GET.get('month', str(timezone.now().month))
    selected_year = request.GET.get('year', str(timezone.now().year))

    late_comers_data = []
    if selected_group_id:
        student_group = get_object_or_404(StudentGroup, pk=selected_group_id)
        late_records = AttendanceRecord.objects.filter(
            timetable__student_group=student_group,
            is_late=True,  # This is the correct filter
            date__month=selected_month,
            date__year=selected_year
        ).values('student__id', 'student__first_name', 'student__last_name').annotate(late_count=Count('student'))

        for record in late_records:
            late_comers_data.append({
                'student_name': f"{record['student__first_name']} {record['student__last_name']}",
                'late_count': record['late_count']
            })

    context = {
        'student_groups': student_groups,
        'selected_group_id': int(selected_group_id) if selected_group_id else None,
        'selected_month': int(selected_month),
        'selected_year': int(selected_year),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'years': range(timezone.now().year, timezone.now().year - 5, -1),
        'late_comers_data': late_comers_data
    }
    return render(request, 'academics/late_comers.html', context)


@login_required
@permission_required('auth.view_user')
def student_profile_view(request, student_id):
    student = get_object_or_404(User, pk=student_id, profile__role='student')

    # --- Initialize all context variables ---
    overall_attendance, total_subjects, current_semester = 0, 0, None
    overall_marks_percentage = 0
    marks_summary = {}

    student_group = student.profile.student_group
    if student_group:
        latest_semester_instance = CourseSubject.objects.filter(
            course=student_group.course
        ).order_by('-semester').first()

        if latest_semester_instance:
            current_semester = latest_semester_instance.semester
            subjects_for_semester = CourseSubject.objects.filter(
                course=student_group.course, semester=current_semester
            )
            total_subjects = subjects_for_semester.count()

            # --- Attendance Calculation ---
            total_classes = AttendanceRecord.objects.filter(
                timetable__student_group=student_group,
                timetable__subject__in=subjects_for_semester
            ).values('date', 'timetable__time_slot').distinct().count()
            present_count = AttendanceRecord.objects.filter(
                student=student,
                timetable__subject__in=subjects_for_semester,
                status__in=['Present', 'Late']
            ).count()
            if total_classes > 0:
                overall_attendance = round((present_count / total_classes) * 100, 1)

            # --- Marks Calculation ---
            marks_qs = Mark.objects.filter(student=student, subject__in=subjects_for_semester)
            total_marks_obtained = marks_qs.aggregate(total=Sum('marks_obtained'))['total'] or 0
            total_max_marks = marks_qs.aggregate(total=Sum('criterion__max_marks'))['total'] or 0
            if total_max_marks > 0:
                overall_marks_percentage = round((total_marks_obtained / total_max_marks) * 100, 1)

            # --- Structure Marks Data for the Template ---
            for mark in marks_qs.select_related('subject__subject', 'criterion').order_by('subject__subject__name'):
                subject_name = mark.subject.subject.name
                if subject_name not in marks_summary:
                    marks_summary[subject_name] = {'criteria': [], 'total_obtained': 0, 'total_max': 0}

                marks_summary[subject_name]['criteria'].append({
                    'name': mark.criterion.name, 'marks': mark.marks_obtained, 'max': mark.criterion.max_marks
                })
                marks_summary[subject_name]['total_obtained'] += mark.marks_obtained
                marks_summary[subject_name]['total_max'] += mark.criterion.max_marks

    context = {
        'student': student, 'overall_attendance': overall_attendance, 'total_subjects': total_subjects,
        'current_semester': current_semester, 'overall_marks_percentage': overall_marks_percentage,
        'marks_summary': marks_summary
    }
    return render(request, 'academics/student_profile.html', context)


@login_required
@permission_required('academics.view_markingscheme')
@nav_item(title="Marking Schemes", icon="iconsminds-testimonal", url_name="academics:scheme_list",
          permission='academics.view_markingscheme', group='admin_management', order=50)
def scheme_list_view(request):
    schemes = MarkingScheme.objects.all()
    return render(request, 'academics/scheme_list.html', {'schemes': schemes})


@login_required
@permission_required('academics.add_markingscheme')
def scheme_create_view(request):
    if request.method == 'POST':
        form = MarkingSchemeForm(request.POST)
        formset = CriterionFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            scheme = form.save()
            formset.instance = scheme
            formset.save()
            messages.success(request, 'Marking scheme created successfully.')
            return redirect('academics:scheme_list')
    else:
        form = MarkingSchemeForm()
        formset = CriterionFormSet()

    context = {
        'form': form,
        'formset': formset,
        'form_title': 'Create New Marking Scheme'
    }
    return render(request, 'academics/scheme_form.html', context)


@login_required
@permission_required('academics.change_markingscheme')
def scheme_update_view(request, pk):
    scheme = get_object_or_404(MarkingScheme, pk=pk)
    if request.method == 'POST':
        form = MarkingSchemeForm(request.POST, instance=scheme)
        formset = CriterionFormSet(request.POST, instance=scheme)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Marking scheme updated successfully.')
            return redirect('academics:scheme_list')
    else:
        form = MarkingSchemeForm(instance=scheme)
        formset = CriterionFormSet(instance=scheme)

    context = {
        'form': form,
        'formset': formset,
        'form_title': f'Edit Marking Scheme: {scheme.name}'
    }
    return render(request, 'academics/scheme_form.html', context)


@login_required
@permission_required('academics.delete_markingscheme')
def scheme_delete_view(request, pk):
    scheme = get_object_or_404(MarkingScheme, pk=pk)
    if request.method == 'POST':
        scheme.delete()
        messages.success(request, 'Marking scheme has been deleted.')
        return redirect('academics:scheme_list')
    return render(request, 'academics/confirm_delete.html', {'item': scheme, 'type': 'Marking Scheme'})


@login_required
@permission_required('academics.add_mark')
@nav_item(title="Enter Marks", icon="simple-icon-note", url_name="academics:marks_entry",
          permission='academics.add_mark', group='faculty_tools', order=30)
def marks_entry_view(request):
    # Get all unique classes the faculty teaches
    taught_entries = Timetable.objects.filter(faculty=request.user).select_related('student_group',
                                                                                   'subject__subject').distinct()
    group_ids = taught_entries.values_list('student_group_id', flat=True).distinct()
    groups_queryset = StudentGroup.objects.filter(id__in=group_ids)

    # Initialize the form with an empty subjects queryset, as it will be populated dynamically
    subjects_queryset = CourseSubject.objects.none()
    form = MarkSelectForm(request.GET or None, groups_queryset=groups_queryset, subjects_queryset=subjects_queryset)

    # Create a map of groups to their current semester's subjects for the dynamic dropdown
    group_subject_map = {}
    # Eager load the course to avoid extra queries inside the loop
    for group in groups_queryset.select_related('course'):
        # Find the latest (current) semester for the group's course
        latest_semester_num = CourseSubject.objects.filter(
            course=group.course
        ).order_by('-semester').values_list('semester', flat=True).first()

        # If no semester is defined for the course, there are no subjects to show
        if not latest_semester_num:
            group_subject_map[group.id] = []
            continue

        # Get all subject IDs the faculty teaches for this specific group - THIS IS THE KEY FIX
        subject_ids = taught_entries.filter(
            student_group=group
        ).values_list('subject_id', flat=True).distinct()

        # Filter the subjects to only include those from the current semester AND taught by this faculty
        current_semester_subjects = CourseSubject.objects.filter(
            id__in=subject_ids,
            semester=latest_semester_num
        ).select_related('subject')

        # Build the map for the dependent dropdown
        group_subject_map[group.id] = [
            {'id': s.id, 'name': s.subject.name} for s in current_semester_subjects
        ]

    students, criteria, existing_marks = None, None, {}
    student_group_id = request.GET.get('student_group')
    course_subject_id = request.GET.get('course_subject')

    if student_group_id and course_subject_id:
        # Verify that the faculty member actually teaches this subject to this class
        if not taught_entries.filter(
                student_group_id=student_group_id,
                subject_id=course_subject_id
        ).exists():
            messages.error(request, "You are not authorized to enter marks for this subject and class combination.")
            return redirect('academics:marks_entry')

        students = User.objects.filter(profile__student_group_id=student_group_id, profile__role='student').order_by(
            'first_name')
        course_subject = get_object_or_404(CourseSubject, pk=course_subject_id)
        active_scheme = course_subject.course.marking_scheme

        if active_scheme:
            criteria = active_scheme.criteria.all()
            marks_qs = Mark.objects.filter(
                student__in=students,
                subject=course_subject,
                criterion__in=criteria
            )
            # Create a nested dictionary for easy lookup in the template
            for mark in marks_qs:
                if mark.student_id not in existing_marks:
                    existing_marks[mark.student_id] = {}
                existing_marks[mark.student_id][mark.criterion_id] = mark.marks_obtained

    if request.method == 'POST':
        # Additional verification for POST requests
        if not taught_entries.filter(
                student_group_id=student_group_id,
                subject_id=course_subject_id
        ).exists():
            messages.error(request, "You are not authorized to enter marks for this subject and class combination.")
            return redirect('academics:marks_entry')

        subject = get_object_or_404(CourseSubject, pk=course_subject_id)
        active_scheme = subject.course.marking_scheme
        if active_scheme and students:
            for student in students:
                for criterion in active_scheme.criteria.all():
                    input_name = f'marks-{student.id}-{criterion.id}'
                    marks_val = request.POST.get(input_name)
                    if marks_val:
                        Mark.objects.update_or_create(
                            student=student, subject=subject, criterion=criterion,
                            defaults={'marks_obtained': marks_val}
                        )
            messages.success(request, "Marks have been saved successfully.")
            return redirect(request.get_full_path())

    context = {
        'form': form,
        'students': students,
        'criteria': criteria,
        'existing_marks': existing_marks,
        'selected_group': student_group_id,
        'selected_subject': course_subject_id,
        'group_subject_map_json': json.dumps(group_subject_map)
    }
    return render(request, 'academics/marks_entry_form.html', context)


@login_required
@permission_required('academics.add_mark')
def download_marks_template_view(request):
    """
    Generates and serves a blank CSV template for bulk marks import.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="marks_import_template.csv"'

    writer = csv.writer(response)
    # These are the required headers for the import to work
    writer.writerow(['student_username', 'subject_code', 'criterion_name', 'marks_obtained'])

    return response


@login_required
@permission_required('academics.add_mark')
@nav_item(title="Bulk Import Marks", icon="iconsminds-up", url_name="academics:bulk_marks_import",
          permission='academics.add_mark', group='faculty_tools')
def bulk_marks_import_view(request):
    if request.method == 'POST':
        form = BulkMarksImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)

            success_count = 0
            error_list = []

            try:
                with transaction.atomic():
                    for row_num, row in enumerate(reader, 2):
                        username = row.get('student_username')
                        subject_code = row.get('subject_code')
                        criterion_name = row.get('criterion_name')
                        marks = row.get('marks_obtained')

                        try:
                            student = User.objects.get(username=username)
                            subject = CourseSubject.objects.get(subject__code=subject_code)
                            criterion = Criterion.objects.get(name=criterion_name)

                            Mark.objects.update_or_create(
                                student=student,
                                subject=subject,
                                criterion=criterion,
                                defaults={'marks_obtained': marks}
                            )
                            success_count += 1

                        except User.DoesNotExist:
                            error_list.append(f"Row {row_num}: Student '{username}' not found.")
                        except CourseSubject.DoesNotExist:
                            error_list.append(f"Row {row_num}: Subject with code '{subject_code}' not found.")
                        except Criterion.DoesNotExist:
                            error_list.append(f"Row {row_num}: Criterion '{criterion_name}' not found.")
                        except Exception as e:
                            error_list.append(f"Row {row_num}: An unexpected error occurred - {e}")

                    if error_list:
                        raise Exception("Import failed, rolling back transaction.")

            except Exception as e:
                messages.error(request, str(e))

            if not error_list:
                messages.success(request, f"Successfully imported marks for {success_count} records.")
            else:
                messages.warning(request,
                                 f"Import completed with errors. Successfully imported: {success_count}. Failed: {len(error_list)}.")

            return render(request, 'academics/bulk_marks_import.html',
                          {'form': BulkMarksImportForm(), 'errors': error_list})
    else:
        form = BulkMarksImportForm()

    return render(request, 'academics/bulk_marks_import.html', {'form': form})


@login_required
@permission_required('academics.view_own_marks')  # Assuming this permission
@nav_item(title="My Marks", icon="simple-icon-graduation", url_name="academics:student_my_marks",
          permission='academics.view_own_marks', group='my_academics', order=3)
def student_my_marks_view(request):
    student = request.user
    settings = AttendanceSettings.load()
    passing_percentage = settings.passing_percentage

    # Get available semesters for the student's course
    available_semesters = []
    if student.profile.student_group:
        available_semesters = CourseSubject.objects.filter(
            course=student.profile.student_group.course
        ).values_list('semester', flat=True).distinct().order_by('semester')

    # Get the selected semester from the request, or default to the latest one
    selected_semester = request.GET.get('semester')
    if not selected_semester and available_semesters:
        selected_semester = available_semesters.last()

    marks_data = {}
    if selected_semester:
        selected_semester = int(selected_semester)
        # Fetch all marks for the student in the selected semester
        marks_qs = Mark.objects.filter(
            student=student,
            subject__semester=selected_semester
        ).select_related('subject__subject', 'criterion')

        # Process the marks into a dictionary grouped by subject
        for mark in marks_qs:
            subject_name = mark.subject.subject.name
            if subject_name not in marks_data:
                marks_data[subject_name] = {
                    'criteria': [],
                    'total_marks': 0,
                    'max_total': 0
                }

            marks_data[subject_name]['criteria'].append({
                'name': mark.criterion.name,
                'marks_obtained': mark.marks_obtained,
                'max_marks': mark.criterion.max_marks
            })
            marks_data[subject_name]['total_marks'] += mark.marks_obtained
            marks_data[subject_name]['max_total'] += mark.criterion.max_marks

            for subject, data in marks_data.items():
                for criterion in data['criteria']:
                    # Calculate if the student passed this criterion
                    passing_marks = (passing_percentage / 100) * criterion['max_marks']
                    criterion['passed'] = criterion['marks_obtained'] >= passing_marks

    context = {
        'marks_data': marks_data,
        'available_semesters': available_semesters,
        'selected_semester': selected_semester,
        'passing_percentage': passing_percentage,
    }
    return render(request, 'academics/student_my_marks.html', context)


@login_required
@permission_required('academics.view_mark')
@nav_item(title="Marks Reports", icon="simple-icon-printer", url_name="academics:marks_report",
          permission='academics.view_mark', group='admin_management')
def marks_report_view(request):
    """
    Displays a form to select criteria for the marks report.
    """
    form = MarksReportForm()
    return render(request, 'academics/marks_report_form.html', {'form': form})


@login_required
@permission_required('academics.view_mark')
def download_marks_report_view(request):
    """
    Generates and serves an Excel file containing the marks for the selected class and semester.
    """
    group_id = request.GET.get('student_group')
    semester = request.GET.get('semester')

    if not group_id or not semester:
        return HttpResponse("Please select a class and semester.", status=400)

    student_group = get_object_or_404(StudentGroup, pk=group_id)
    students = User.objects.filter(profile__student_group=student_group).order_by('first_name')
    course_subjects = CourseSubject.objects.filter(
        course=student_group.course,
        semester=semester
    ).select_related('subject')

    # Create an in-memory workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Sem {semester} Marks - {student_group.name}"

    # --- Create Header Row ---
    headers = ["Student Name", "Student ID"] + [cs.subject.name for cs in course_subjects]
    sheet.append(headers)
    for cell in sheet[1]:  # Style the header row
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Prepare data ---
    # Get all marks for the selected students and subjects in one query
    all_marks = Mark.objects.filter(
        student__in=students,
        subject__in=course_subjects
    ).values(
        'student_id', 'subject_id'
    ).annotate(
        total_marks=Sum('marks_obtained')
    )

    # Pivot the data for easy lookup
    marks_pivot = {}
    for mark in all_marks:
        if mark['student_id'] not in marks_pivot:
            marks_pivot[mark['student_id']] = {}
        marks_pivot[mark['student_id']][mark['subject_id']] = mark['total_marks']

    # --- Populate Data Rows ---
    for student in students:
        row_data = [student.get_full_name(), student.profile.student_id_number]
        student_marks = marks_pivot.get(student.id, {})
        for cs in course_subjects:
            marks = student_marks.get(cs.id, 0)  # Default to 0 if no mark found
            row_data.append(marks)
        sheet.append(row_data)

    # --- Prepare the response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Marks_{student_group.name}_Sem_{semester}.xlsx"'
    workbook.save(response)

    return response


@login_required
@permission_required('accounts.view_own_profile')  # Assuming a base permission
@nav_item(title="My Profile", icon="simple-icon-user", url_name="academics:my_profile",
          permission='accounts.view_own_profile', group='my_academics', order=0)
def my_profile_view(request):
    """
    Displays the profile page for the currently logged-in user.
    """
    student = request.user

    # --- FIX STARTS HERE: Added full logic to fetch profile data ---
    overall_attendance, total_subjects, current_semester = 0, 0, None
    overall_marks_percentage = 0
    marks_summary = {}

    # Ensure the user is a student with a profile before calculating
    if hasattr(student, 'profile') and student.profile.role == 'student':
        student_group = student.profile.student_group
        if student_group:
            latest_semester_instance = CourseSubject.objects.filter(
                course=student_group.course
            ).order_by('-semester').first()

            if latest_semester_instance:
                current_semester = latest_semester_instance.semester
                subjects_for_semester = CourseSubject.objects.filter(
                    course=student_group.course, semester=current_semester
                )
                total_subjects = subjects_for_semester.count()

                # --- Attendance Calculation ---
                total_classes = AttendanceRecord.objects.filter(
                    timetable__student_group=student_group,
                    timetable__subject__in=subjects_for_semester
                ).values('date', 'timetable__time_slot').distinct().count()
                present_count = AttendanceRecord.objects.filter(
                    student=student,
                    timetable__subject__in=subjects_for_semester,
                    status__in=['Present', 'Late']
                ).count()
                if total_classes > 0:
                    overall_attendance = round((present_count / total_classes) * 100, 1)

                # --- Marks Calculation ---
                marks_qs = Mark.objects.filter(student=student, subject__in=subjects_for_semester)
                total_marks_obtained = marks_qs.aggregate(total=Sum('marks_obtained'))['total'] or 0
                total_max_marks = marks_qs.aggregate(total=Sum('criterion__max_marks'))['total'] or 0
                if total_max_marks > 0:
                    overall_marks_percentage = round((total_marks_obtained / total_max_marks) * 100, 1)

                # --- Structure Marks Data for the Template ---
                for mark in marks_qs.select_related('subject__subject', 'criterion').order_by('subject__subject__name'):
                    subject_name = mark.subject.subject.name
                    if subject_name not in marks_summary:
                        marks_summary[subject_name] = {'criteria': [], 'total_obtained': 0, 'total_max': 0}

                    marks_summary[subject_name]['criteria'].append({
                        'name': mark.criterion.name, 'marks': mark.marks_obtained, 'max': mark.criterion.max_marks
                    })
                    marks_summary[subject_name]['total_obtained'] += mark.marks_obtained
                    marks_summary[subject_name]['total_max'] += mark.criterion.max_marks

    context = {
        'student': student,
        'overall_attendance': overall_attendance,
        'total_subjects': total_subjects,
        'current_semester': current_semester,
        'overall_marks_percentage': overall_marks_percentage,
        'marks_summary': marks_summary
    }
    # --- FIX ENDS HERE ---

    # We can reuse the same template as the admin-facing profile view
    return render(request, 'academics/student_profile.html', context)


@login_required
@permission_required('academics.add_extraclass')
@nav_item(title="Schedule Extra Class", icon="simple-icon-plus", url_name="academics:schedule_extra_class",
          permission='academics.add_extraclass', group='faculty_tools', order=40)
def schedule_extra_class(request):
    if request.method == 'POST':
        form = ExtraClassForm(request.POST, user=request.user)
        if form.is_valid():
            with transaction.atomic():
                # Your existing logic to save the class and create an announcement
                extra_class = form.save(commit=False)
                if request.user.profile.role == 'faculty':
                    extra_class.teacher = request.user
                extra_class.save()

                announcement = Announcement.objects.create(
                    title=f"Extra Class: {extra_class.subject.subject.name}",
                    content=f"An extra class for {extra_class.subject.subject.name} has been scheduled on {extra_class.date} during the {extra_class.time_slot} slot.",
                    sender=request.user
                )
                announcement.target_student_groups.add(extra_class.class_group)
                extra_class.announcement = announcement
                extra_class.save()

                # === NEW: AUTOMATIC EMAIL NOTIFICATION LOGIC ===
                try:
                    student_group = extra_class.class_group
                    students = User.objects.filter(profile__student_group=student_group)
                    recipient_emails = [student.email for student in students if student.email]

                    if recipient_emails:
                        subject = f"Extra Class Scheduled: {extra_class.subject.subject.name} on {extra_class.date.strftime('%d-%b-%Y')}"
                        email_context = {'extra_class': extra_class}

                        # Render HTML and plain text versions of the email
                        html_content = render_to_string('emails/extra_class_notification.html', email_context)
                        plain_text_content = (
                            f"Dear Student,\n\nAn extra class has been scheduled.\n\n"
                            f"Subject: {extra_class.subject.subject.name}\n"
                            f"Faculty: {extra_class.teacher.get_full_name()}\n"
                            f"Date: {extra_class.date.strftime('%A, %B %d, %Y')}\n"
                            f"Time: {extra_class.time_slot.start_time.strftime('%I:%M %p')} - {extra_class.time_slot.end_time.strftime('%I:%M %p')}\n\n"
                            f"Please ensure your attendance."
                        )

                        # Send one email to all students via BCC for privacy
                        send_database_email(
                            subject=subject,
                            body=plain_text_content,
                            recipient_list=[AttendanceSettings.load().email_host_user],
                            html_message=html_content,
                            bcc_list=recipient_emails
                        )
                        messages.success(request,
                                         f"Extra class scheduled and a notification email has been sent to {len(recipient_emails)} students.")
                    else:
                        messages.success(request,
                                         'Extra class scheduled successfully (no students found in the group to notify).')

                except Exception as e:
                    messages.error(request,
                                   f"Extra class was scheduled, but the notification email failed to send. Error: {e}")
                # === END OF NOTIFICATION LOGIC ===

            return redirect('academics:extra_class_list')
        else:
            messages.warning(request, "Please complete all required fields.")
    else:
        initial_data = {}
        if 'teacher' in request.GET:
            initial_data['teacher'] = request.GET.get('teacher')
        form = ExtraClassForm(user=request.user, initial=initial_data)

    context = {
        'form': form,
        'form_title': 'Schedule an Extra Class'
    }
    return render(request, 'academics/extra_class_form.html', context)


@login_required
@permission_required('academics.update_extraclass')
def extra_class_update(request, pk):
    extra_class = get_object_or_404(ExtraClass, pk=pk)
    if request.method == 'POST':
        form = ExtraClassForm(request.POST, instance=extra_class, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Extra class updated successfully.')
            return redirect('academics:extra_class_list')
    else:
        # --- FIX: Handle GET requests for filtering ---
        initial_data = {}
        # If a teacher was selected via the dropdown reload, pass it to the form
        if 'teacher' in request.GET:
            initial_data['teacher'] = request.GET.get('teacher')

        form = ExtraClassForm(instance=extra_class, user=request.user, initial=initial_data)
        # --- END FIX ---

    return render(request, 'academics/extra_class_form.html', {'form': form, 'form_title': f'Edit Extra Class'})


@login_required
@permission_required('academics.view_extraclass')
@nav_item(title="Extra Class", icon="simple-icon-clock", url_name="academics:extra_class_list",
          permission='accounts.view_extraclass', group='admin_management', order=0)
def extra_class_list(request):
    extra_classes = ExtraClass.objects.all()
    return render(request, 'academics/extra_class_list.html', {'extra_classes': extra_classes})


@login_required
@permission_required('academics.delete_extraclass')
def extra_class_delete(request, pk):
    extra_class = get_object_or_404(ExtraClass, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            # --- Step 1: Send Notifications FIRST ---
            try:
                # Get the list of students before the class is deleted
                students = User.objects.filter(profile__student_group=extra_class.class_group)
                recipient_emails = [student.email for student in students if student.email]

                if recipient_emails:
                    subject = f"Class Cancelled: {extra_class.subject.subject.name} on {extra_class.date.strftime('%d-%b-%Y')}"
                    email_context = {'extra_class': extra_class}

                    # Render the cancellation email templates
                    html_content = render_to_string('emails/extra_class_cancelled_notification.html', email_context)
                    plain_text_content = f"The extra class for {extra_class.subject.subject.name} on {extra_class.date} has been cancelled."

                    # Use our utility to send the email via BCC
                    send_database_email(
                        subject=subject,
                        body=plain_text_content,
                        recipient_list=[AttendanceSettings.load().email_host_user],
                        html_message=html_content,
                        bcc_list=recipient_emails
                    )
            except Exception as e:
                # If email fails, log it and inform the admin, but don't stop the process
                messages.error(request, f"Could not send cancellation emails. Error: {e}")

            # --- Step 2: Create the on-site announcement ---
            cancellation_announcement = Announcement.objects.create(
                title=f"CANCELLED: Extra Class for {extra_class.subject}",
                content=f"The extra class for {extra_class.subject} scheduled on {extra_class.date} has been cancelled.",
                sender=request.user
            )
            cancellation_announcement.target_student_groups.add(extra_class.class_group)

            # --- Step 3: Delete the record LAST ---
            extra_class.delete()

        messages.success(request, 'Extra class has been cancelled and students have been notified.')
        return redirect('academics:extra_class_list')

    # This part for the GET request remains the same
    context = {
        'item': extra_class,
        'type': 'Extra Class'
    }
    return render(request, 'academics/confirm_delete.html', context)


@login_required
@nav_item(title="User Guide", icon="simple-icon-question", url_name="academics:guide",
          permission=None, order=10)
def guide_view(request):
    """
    Renders the user guide, passing the user's role and the
    currently selected section to the template.
    """
    user_profile = request.user.profile

    # Determine the role from the user's profile
    # This matches the strings used in the guide.html template
    user_role = user_profile.role

    # Get the requested section from the URL (?section=...), defaulting to 'getting-started'
    current_section = request.GET.get('section', 'getting-started')
    print(user_role)

    context = {
        'page_title': 'User Guide',
        'user_role': user_role,
        'current_section': current_section
    }
    return render(request, 'academics/guide.html', context)


@login_required
def student_report_card_html_view(request, student_id):
    student = get_object_or_404(User, pk=student_id, profile__role='student')
    student_group = student.profile.student_group

    context_data = {'student': student, 'student_group': student_group}

    if student_group:
        # Get ALL semesters instead of just the latest one
        all_semesters = CourseSubject.objects.filter(
            course=student_group.course
        ).values_list('semester', flat=True).distinct().order_by('semester')

        # Create a dictionary to store performance data by semester
        semester_performance = {}
        overall_totals = {
            'total_held': 0,
            'total_attended': 0,
            'total_marks_obtained': 0,
            'total_max_marks': 0
        }

        for semester in all_semesters:
            subjects_for_semester = CourseSubject.objects.filter(
                course=student_group.course,
                semester=semester
            )

            # Get attendance and marks for this semester
            all_records = AttendanceRecord.objects.filter(student=student).filter(
                Q(timetable__subject__in=subjects_for_semester) |
                Q(extra_class__subject__in=subjects_for_semester)
            )
            all_marks = Mark.objects.filter(student=student, subject__in=subjects_for_semester)

            performance_data = []
            semester_held = 0
            semester_attended = 0
            semester_marks_obtained = 0
            semester_max_marks = 0

            for cs in subjects_for_semester:
                subject_total_held = AttendanceRecord.objects.filter(
                    (Q(timetable__student_group=student_group) & Q(timetable__subject=cs)) |
                    (Q(extra_class__class_group=student_group) & Q(extra_class__subject=cs))
                ).values('date', 'timetable_id', 'extra_class_id').distinct().count()

                subject_attended = all_records.filter(
                    (Q(timetable__subject=cs) | Q(extra_class__subject=cs)) &
                    Q(status__in=['Present', 'Late'])
                ).count()

                subject_marks = all_marks.filter(subject=cs)
                subject_marks_obtained = subject_marks.aggregate(total=Sum('marks_obtained'))['total'] or 0
                subject_max_marks = subject_marks.aggregate(total=Sum('criterion__max_marks'))['total'] or 0

                performance_data.append({
                    'subject_name': cs.subject.name,
                    'attendance_percentage': round((subject_attended / subject_total_held) * 100,
                                                   1) if subject_total_held > 0 else 0,
                    'marks_obtained': subject_marks_obtained,
                    'max_marks': subject_max_marks
                })

                # Add to semester totals
                semester_held += subject_total_held
                semester_attended += subject_attended
                semester_marks_obtained += subject_marks_obtained
                semester_max_marks += subject_max_marks

            # Calculate semester percentages
            semester_attendance_percentage = round((semester_attended / semester_held) * 100,
                                                   1) if semester_held > 0 else 0
            semester_marks_percentage = round((semester_marks_obtained / semester_max_marks) * 100,
                                              1) if semester_max_marks > 0 else 0

            semester_performance[semester] = {
                'performance_data': performance_data,
                'semester_attendance': semester_attendance_percentage,
                'semester_marks_obtained': semester_marks_obtained,
                'semester_max_marks': semester_max_marks,
                'semester_marks_percentage': semester_marks_percentage
            }

            # Add to overall totals
            overall_totals['total_held'] += semester_held
            overall_totals['total_attended'] += semester_attended
            overall_totals['total_marks_obtained'] += semester_marks_obtained
            overall_totals['total_max_marks'] += semester_max_marks

        # Calculate overall percentages
        overall_attendance = round((overall_totals['total_attended'] / overall_totals['total_held']) * 100, 1) if \
            overall_totals['total_held'] > 0 else 0
        overall_marks_percentage = round(
            (overall_totals['total_marks_obtained'] / overall_totals['total_max_marks']) * 100, 1) if overall_totals[
                                                                                                          'total_max_marks'] > 0 else 0

        context_data.update({
            'semester_performance': semester_performance,
            'all_semesters': all_semesters,
            'overall_attendance': overall_attendance,
            'overall_marks': overall_totals['total_marks_obtained'],
            'overall_max_marks': overall_totals['total_max_marks'],
            'overall_marks_percentage': overall_marks_percentage,
        })

    return render(request, 'academics/report_card.html', context_data)


# --- VIEW 2: The ASYNC view that controls the browser and serves the PDF ---
async def student_report_card_pdf_view(request, student_id):
    """
    Generates a PDF report card by launching a headless browser with Pyppeteer,
    visiting the HTML version of the report, and printing it to PDF.
    """
    # This URL must match the one in your urls.py for the HTML view
    render_url = request.build_absolute_uri(reverse('academics:student_report_card_html', args=[student_id]))

    try:
        browser = await launch(headless=True, args=['--no-sandbox'])
        page = await browser.newPage()

        # Go to the page and wait until network traffic has settled
        await page.goto(render_url, {'waitUntil': 'networkidle0'})

        # --- FIX: Added 'displayHeaderFooter': False to remove the default header ---
        pdf_content = await page.pdf({
            'format': 'A4',
            'landscape': True,
            'printBackground': True,
            'displayHeaderFooter': False,  # This removes the unwanted top/bottom text
            'margin': {  # Optional: You can adjust margins for a better fit
                'top': '0.5in',
                'bottom': '0.5in',
                'left': '0.5in',
                'right': '0.5in'
            }
        })

        await browser.close()

        # Serve the PDF as a download
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="report_card_{student_id}.pdf"'

        return response
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {e}", status=500)


async def student_report_card_pdf_view(request, student_id):
    """
    Generates a PDF report card by launching a headless browser with Pyppeteer,
    visiting the HTML version of the report, and printing it to PDF.
    """
    # This URL must match the one in your urls.py for the HTML view
    render_url = request.build_absolute_uri(reverse('academics:student_report_card_html', args=[student_id]))

    try:
        browser = await launch(headless=True, args=['--no-sandbox'])
        page = await browser.newPage()

        # Go to the page and wait until network traffic has settled
        await page.goto(render_url, {'waitUntil': 'networkidle0'})

        # --- FIX: Added 'displayHeaderFooter': False to remove the default header ---
        pdf_content = await page.pdf({
            'format': 'A4',
            'landscape': True,
            'printBackground': True,
            'displayHeaderFooter': False,  # This removes the unwanted top/bottom text
            'margin': {  # Optional: You can adjust margins for a better fit
                'top': '0.5in',
                'bottom': '0.5in',
                'left': '0.5in',
                'right': '0.5in'
            }
        })

        await browser.close()

        # Serve the PDF as a download
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="report_card_{student_id}.pdf"'

        return response
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {e}", status=500)


@login_required
@permission_required('academics.add_extraclass')
def get_teacher_class_subjects_view(request):
    teacher_id = request.GET.get('teacher_id')
    class_group_id = request.GET.get('class_group_id')

    subjects = []
    if teacher_id and class_group_id:
        try:
            teacher = User.objects.get(pk=teacher_id)
            class_group = StudentGroup.objects.get(pk=class_group_id)

            course_subjects = CourseSubject.objects.filter(
                timetable_entries__faculty=teacher,
                timetable_entries__student_group=class_group
            ).select_related('subject').distinct()

            subjects = [
                {'id': cs.id, 'name': cs.subject.name}
                for cs in course_subjects
            ]
        except (User.DoesNotExist, StudentGroup.DoesNotExist):
            pass

    return JsonResponse({'subjects': subjects})


@login_required
def get_subject_faculty_view(request):
    subject_id = request.GET.get('subject_id')
    faculty_data = []

    if subject_id:
        try:
            subject = CourseSubject.objects.get(pk=subject_id)

            # Get specialized faculty first
            specialized_faculty = User.objects.filter(
                profile__role='faculty',
                profile__field_of_expertise=subject.subject
            ).order_by('first_name')

            # Add specialized faculty first
            for faculty in specialized_faculty:
                faculty_data.append({
                    'id': faculty.id,
                    'name': faculty.get_full_name(),
                    'specialized': True
                })


        except CourseSubject.DoesNotExist:
            pass

    return JsonResponse({'faculty': faculty_data})


@login_required
@permission_required('academics.add_backup')
@nav_item(title="Backup & Restore", icon="simple-icon-cloud-download", url_name="academics:backup_restore",
          group="application_settings", order=50, role_required='admin')
def backup_restore_view(request):
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    if request.method == 'POST':
        if 'create_backup' in request.POST:
            try:
                timestamp = timezone.now().strftime('%Y-%m-%d_%H-%M-%S')
                backup_filename = f'database_backup_{timestamp}.json'
                backup_filepath = os.path.join(backup_dir, backup_filename)

                with open(backup_filepath, 'w') as backup_file:
                    call_command('dumpdata',
                                 '--natural-foreign',
                                 '--natural-primary',
                                 '--exclude=contenttypes',
                                 '--exclude=auth.permission',
                                 '--exclude=sessions.session',
                                 '--exclude=admin.logentry',
                                 '--indent=2',
                                 stdout=backup_file)

                # Clean up old backups (keep only last 10)
                backup_files = sorted([f for f in os.listdir(backup_dir) if f.endswith('.json')])
                if len(backup_files) > 10:
                    for old_file in backup_files[:-10]:
                        os.remove(os.path.join(backup_dir, old_file))

                messages.success(request, f"Database backup created successfully: {backup_filename}")
            except Exception as e:
                messages.error(request, f"An error occurred while creating the backup: {e}")

        elif 'restore_backup' in request.POST:
            backup_file = request.POST.get('backup_file')
            if backup_file:
                backup_filepath = os.path.join(backup_dir, backup_file)
                if os.path.exists(backup_filepath):
                    try:
                        from django.contrib.auth import logout
                        from django.http import HttpResponseRedirect
                        from django.db import transaction

                        # Store superuser info before logout
                        superusers = list(User.objects.filter(is_superuser=True).values(
                            'username', 'email', 'first_name', 'last_name', 'password', 'is_staff', 'is_active'
                        ))

                        # Log out the user BEFORE restoration
                        logout(request)

                        # Use the enhanced restore command with clear option
                        with transaction.atomic():
                            # Call the management command with clear flag
                            call_command('manage_backups', 'restore',
                                         '--file', backup_filepath,
                                         '--clear',
                                         '--preserve-superusers')

                        return HttpResponseRedirect(f"{reverse('accounts:login')}?restore_success=1")

                    except Exception as e:
                        messages.error(request, f"An error occurred during restore: {e}")
                        return redirect('academics:backup_restore')
                else:
                    messages.error(request, "The selected backup file was not found.")
            else:
                messages.error(request, "You must select a backup file to restore.")

        return redirect('academics:backup_restore')

    # List available backups
    backup_files = []
    try:
        files = [f for f in os.listdir(backup_dir) if f.endswith('.json')]
        for filename in sorted(files, reverse=True):
            filepath = os.path.join(backup_dir, filename)
            stat = os.stat(filepath)

            try:
                timestamp_str = filename.replace('database_backup_', '').replace('.json', '')
                timestamp_str = timestamp_str.replace('_', ' ').replace('-', ':')
                created_date = timezone.datetime.strptime(timestamp_str.replace(':', '-', 2), '%Y-%m-%d %H-%M-%S')
                formatted_date = created_date.strftime('%B %d, %Y at %I:%M %p')
            except:
                formatted_date = timezone.datetime.fromtimestamp(stat.st_mtime).strftime('%B %d, %Y at %I:%M %p')

            backup_files.append({
                'name': filename,
                'size': round(stat.st_size / (1024 * 1024), 2),
                'created': formatted_date,
                'display_name': filename.replace('database_backup_', '').replace('.json', '').replace('_',
                                                                                                      ' at ').replace(
                    '-', '/')
            })
    except OSError:
        pass

    context = {
        'backup_files': backup_files,
        'page_title': 'Database Backup & Restore'
    }
    return render(request, 'academics/backup_restore.html', context)


@login_required
@permission_required("academics.change_smtp_settings")
@nav_item(title="SMTP Settings", icon="simple-icon-envelope-letter", url_name="academics:smtp_settings",
          permission='academics.change_smtp_settings', group='application_settings', order=60)
def smtp_settings_view(request):
    settings = AttendanceSettings.load()
    if request.method == 'POST':
        form = SmtpSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'SMTP settings have been updated successfully.')
            return redirect('academics:smtp_settings')
    else:
        form = SmtpSettingsForm(instance=settings)

    context = {
        'form': form,
        'page_title': 'SMTP Email Settings'
    }
    return render(request, 'academics/smtp_settings.html', context)


@login_required
@permission_required('academics.view_accesslog')
@nav_item(title="System Reports", icon="simple-icon-chart", url_name="academics:system_reports",
          permission='academics.view_accesslog', group='application_settings', order=70)
def system_reports_view(request):
    # Fetch the last 100 access logs, showing the newest first
    activity_logs = UserActivityLog.objects.all().order_by('-timestamp')[:100]
    log_file_path = os.path.join(settings.BASE_DIR, 'debug.log')
    log_content = []
    try:
        with open(log_file_path, 'r') as f:
            # Read the last 100 lines for efficiency
            log_content = f.readlines()[-100:]
    except FileNotFoundError:
        log_content = ["Log file not found. It will be created when the first log message is written."]

    context = {
        'page_title': 'System Reports',
        'activity_logs': activity_logs,
        'log_content': log_content,
    }
    return render(request, 'academics/system_reports.html', context)


@login_required
@permission_required('academics.view_update_status')
def update_status_view(request):
    update_count = 0
    try:
        # This is the same reliable logic used in the context processor.
        # It directly checks the Git repository for available updates.
        project_dir = '/var/www/AttendanceManagment'
        
        subprocess.run(['git', 'fetch'], check=True, cwd=project_dir)
        
        result = subprocess.run(
            ['git', 'rev-list', '--count', 'HEAD..origin/website-server'],
            capture_output=True, text=True, check=True, cwd=project_dir
        )
        update_count = int(result.stdout.strip())

    except Exception as e:
        logger.error(f"Failed to check for updates on status page: {e}")
        messages.error(request, "Could not check for updates. There might be an issue with Git permissions on the server.")

    context = {
        'page_title': 'Application Update Status',
        'update_count': update_count
    }
    return render(request, 'academics/update_status.html', context)


# In academics/views.py

@login_required
@permission_required('academics.send_bulk_email')
@nav_item(title="Bulk Email", icon="simple-icon-envelope", url_name="academics:bulk_email",
          permission='academics.send_bulk_email', group='admin_management', order=80)
def bulk_email_view(request):
    if request.method == 'POST':
        form = BulkEmailForm(request.POST)
        if form.is_valid():
            recipients = form.cleaned_data['recipients']
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']

            # Prepare the email content
            html_content = render_to_string('emails/custom_bulk_email.html', {'subject': subject, 'message': message})
            from django.utils.html import strip_tags
            plain_text_content = strip_tags(html_content)

            # Get the list of email addresses
            recipient_emails = []
            for r in recipients:
                if r == 'all_students':
                    recipient_emails.extend(
                        User.objects.filter(profile__role='student').values_list('email', flat=True))
                elif r == 'all_faculty':
                    recipient_emails.extend(
                        User.objects.filter(profile__role='faculty').values_list('email', flat=True))
                elif r.startswith('group_'):
                    group_id = int(r.split('_')[1])
                    recipient_emails.extend(
                        User.objects.filter(profile__student_group__id=group_id).values_list('email', flat=True))

            # Remove duplicate emails and filter out any empty strings
            bcc_recipients = list(set(filter(None, recipient_emails)))

            if bcc_recipients:
                # === THIS IS THE NEW, EFFICIENT METHOD ===
                # We send ONE email. The 'To' field can be the sender's own email or a no-reply address.
                # All actual recipients are placed in the BCC list for privacy.
                send_database_email(
                    subject=subject,
                    body=plain_text_content,
                    recipient_list=[AttendanceSettings.load().email_host_user],  # Send "To" the system's own email
                    html_message=html_content,
                    bcc_list=bcc_recipients  # <-- All users are here
                )
                messages.success(request, f"Email successfully queued for sending to {len(bcc_recipients)} recipients.")
                # ==========================================
            else:
                messages.warning(request, "No valid recipients found for the selected groups.")

            return redirect('academics:bulk_email')
    else:
        form = BulkEmailForm()

    context = {
        'form': form,
        'page_title': 'Send Bulk Email'
    }
    return render(request, 'academics/bulk_email.html', context)


@login_required
@permission_required('academics.publish_results')
@nav_item(title="Publish Results", icon="simple-icon-check", url_name="academics:publish_results",
          permission='academics.publish_results', group='admin_management', order=90)
def publish_results_view(request):
    student_groups = StudentGroup.objects.all()
    selected_group = None
    selected_semester = None
    student_statuses = []

    if request.method == 'GET' and 'student_group' in request.GET and 'semester' in request.GET:
        group_id = request.GET.get('student_group')
        semester_str = request.GET.get('semester')
        if group_id and semester_str:
            selected_group = get_object_or_404(StudentGroup, pk=group_id)
            selected_semester = int(semester_str)

            # Get all subjects for the selected course and semester
            subjects_in_semester = CourseSubject.objects.filter(
                course=selected_group.course, semester=selected_semester
            )
            total_subjects_count = subjects_in_semester.count()

            # Get all students in the group
            students_in_group = User.objects.filter(profile__student_group=selected_group)

            for student in students_in_group:
                # Count how many marks have been entered for this student for this semester's subjects
                marks_entered_count = Mark.objects.filter(
                    student=student,
                    subject__in=subjects_in_semester
                ).count()

                # Check if results have already been published for this student/semester
                is_published = ResultPublication.objects.filter(
                    student=student,
                    student_group=selected_group,
                    semester=selected_semester
                ).exists()

                # Determine status
                all_marks_entered = marks_entered_count >= total_subjects_count

                student_statuses.append({
                    'student': student,
                    'all_marks_entered': all_marks_entered,
                    'is_published': is_published,
                    'parent_email': student.profile.parent_email
                })

    context = {
        'page_title': 'Publish Student Results',
        'student_groups': student_groups,
        'selected_group': selected_group,
        'selected_semester': selected_semester,
        'student_statuses': student_statuses,
    }
    return render(request, 'academics/publish_results.html', context)


@login_required
@permission_required('academics.publish_results')
def send_parent_report_email_view(request, student_id, group_id, semester):
    if request.method == 'POST':
        student = get_object_or_404(User, pk=student_id)
        student_group = get_object_or_404(StudentGroup, pk=group_id)
        parent_email = student.profile.parent_email

        if not parent_email:
            messages.error(request, f"No parent email found for {student.get_full_name()}.")
            return redirect(request.META.get('HTTP_REFERER'))

        # Check if already published to prevent duplicate sending
        if ResultPublication.objects.filter(student=student, student_group=student_group, semester=semester).exists():
            messages.warning(request, f"Results for {student.get_full_name()} have already been published.")
            return redirect(request.META.get('HTTP_REFERER'))

        # Get all marks for the report card
        subjects_in_semester = CourseSubject.objects.filter(course=student_group.course, semester=semester)
        marks = Mark.objects.filter(student=student, subject__in=subjects_in_semester)
        settings = AttendanceSettings.load()

        # Prepare email content
        subject = f"Final Report Card for {student.get_full_name()} - Semester {semester}"
        email_context = {
            'student': student,
            'student_group': student_group,
            'semester': semester,
            'marks': marks,
            'passing_percentage': settings.passing_percentage,
        }
        html_content = render_to_string('emails/parent_report_card_email.html', email_context)
        plain_text_content = f"Please find the attached report card for {student.get_full_name()}."  # Fallback text

        # Send the email
        success = send_database_email(
            subject=subject,
            body=plain_text_content,
            recipient_list=[parent_email],
            html_message=html_content
        )

        if success:
            # Log that this report has been published
            ResultPublication.objects.create(
                student=student,
                student_group=student_group,
                semester=semester,
                published_by=request.user
            )
            messages.success(request, f"Report card successfully sent to parent of {student.get_full_name()}.")
        else:
            messages.error(request, f"Failed to send email for {student.get_full_name()}. Please check SMTP settings.")

        return redirect(request.META.get('HTTP_REFERER'))

    return redirect('academics:publish_results')


@login_required
@permission_required('academics.publish_results')
def bulk_publish_results_view(request, group_id, semester):
    if request.method == 'POST':
        student_group = get_object_or_404(StudentGroup, pk=group_id)

        # Get all subjects for the semester
        subjects_in_semester = CourseSubject.objects.filter(course=student_group.course, semester=semester)
        # Get all possible grading criteria for this course's marking scheme
        all_criteria = Criterion.objects.filter(scheme=student_group.course.marking_scheme).order_by('id')

        students_in_group = User.objects.filter(profile__student_group=student_group)
        settings = AttendanceSettings.load()

        published_count = 0
        error_count = 0

        for student in students_in_group:
            # --- Check eligibility for this student ---
            is_already_published = ResultPublication.objects.filter(student=student, student_group=student_group,
                                                                    semester=semester).exists()
            marks_entered_count = Mark.objects.filter(student=student, subject__in=subjects_in_semester).count()

            # This check ensures we have a mark for every subject-criterion pair
            required_marks_count = subjects_in_semester.count() * all_criteria.count()

            if marks_entered_count >= required_marks_count and not is_already_published and student.profile.parent_email:
                # --- DYNAMICALLY PIVOT THE DATA FOR THE EMAIL TEMPLATE ---
                all_marks = Mark.objects.filter(student=student, subject__in=subjects_in_semester).select_related(
                    'subject__subject', 'criterion')

                # Create a nested dictionary to hold the pivoted data: {subject_name: {criterion_name: marks, ...}, ...}
                report_data = {s.subject.name: {} for s in subjects_in_semester}

                for mark in all_marks:
                    subject_name = mark.subject.subject.name
                    criterion_name = mark.criterion.name
                    report_data[subject_name][criterion_name] = {
                        'obtained': mark.marks_obtained,
                        'max': mark.criterion.max_marks
                    }

                # Calculate totals and pass/fail status for each subject
                final_results = []
                for subject_name, criteria_marks in report_data.items():
                    total_obtained = sum(v['obtained'] for v in criteria_marks.values())
                    total_max = sum(v['max'] for v in criteria_marks.values())
                    percentage = (total_obtained / total_max * 100) if total_max > 0 else 0
                    status = "Pass" if percentage >= settings.passing_percentage else "Fail"

                    final_results.append({
                        'subject': subject_name,
                        'criteria_marks': criteria_marks,  # Pass the detailed breakdown
                        'total_obtained': total_obtained,
                        'total_max': total_max,
                        'status': status
                    })
                # --- END OF DYNAMIC PIVOT LOGIC ---

                subject = f"Final Report Card for {student.get_full_name()} - Semester {semester}"
                email_context = {
                    'student': student, 'student_group': student_group, 'semester': semester,
                    'all_criteria': all_criteria,  # Pass the list of criteria for the headers
                    'final_results': final_results,
                }
                html_content = render_to_string('emails/parent_report_card_email.html', email_context)
                plain_text_content = f"Please find the attached report card for {student.get_full_name()}."

                success = send_database_email(subject, plain_text_content, [student.profile.parent_email],
                                              html_message=html_content)

                if success:
                    ResultPublication.objects.create(
                        student=student, student_group=student_group,
                        semester=semester, published_by=request.user
                    )
                    published_count += 1
                else:
                    error_count += 1

        # --- Provide summary feedback to the admin ---
        if published_count > 0:
            messages.success(request, f"Successfully published and sent {published_count} report cards.")
        if error_count > 0:
            messages.error(request, f"Failed to send {error_count} report cards. Please check SMTP settings.")
        if published_count == 0 and error_count == 0:
            messages.info(request, "No new eligible students found to publish results for at this time.")

        return redirect('academics:publish_results')

    return redirect('academics:publish_results')


@login_required
@permission_required('academics.finalize_results')
@nav_item(title="Finalize Results", icon="simple-icon-check", url_name="academics:finalize_results",
          permission='academics.finalize_results', group='admin_management', order=100)
def finalize_results_view(request):
    if request.method == 'POST':
        group_id = request.POST.get('student_group')
        semester_str = request.POST.get('semester')

        if group_id and semester_str:
            student_group = get_object_or_404(StudentGroup, pk=group_id)
            semester = int(semester_str)
            settings = AttendanceSettings.load()

            # Get all subjects and students for the selected group/semester
            subjects_in_semester = CourseSubject.objects.filter(course=student_group.course, semester=semester)
            students_in_group = User.objects.filter(profile__student_group=student_group)

            finalized_count = 0
            for student in students_in_group:
                for subject in subjects_in_semester:
                    # Check if a status for this subject has already been finalized to avoid duplicates
                    is_finalized = StudentSubjectStatus.objects.filter(student=student, subject=subject,
                                                                       semester=semester).exists()
                    if is_finalized:
                        continue

                    # Calculate the total marks and percentage for this subject
                    marks_for_subject = Mark.objects.filter(student=student, subject=subject)
                    total_obtained = sum(m.marks_obtained for m in marks_for_subject)
                    total_max = sum(m.criterion.max_marks for m in marks_for_subject)
                    percentage = (total_obtained / total_max * 100) if total_max > 0 else 0

                    # Determine the final status
                    final_status = 'PASSED' if percentage >= settings.passing_percentage else 'FAILED'

                    # Create the final status record
                    StudentSubjectStatus.objects.create(
                        student=student,
                        subject=subject,
                        semester=semester,
                        status=final_status
                    )
                    finalized_count += 1

            if finalized_count > 0:
                messages.success(request,
                                 f"Successfully finalized results for {finalized_count} student-subject records.")
            else:
                messages.info(request, "All results for this group/semester have already been finalized.")

        return redirect('academics:finalize_results')

    # For a GET request, get student groups from the current session to populate the dropdown
    student_groups = StudentGroup.objects.all()
    context = {
        'page_title': 'Finalize Semester Results',
        'student_groups': student_groups,
    }
    return render(request, 'academics/finalize_results.html', context)


@login_required
@permission_required('academics.manage_supplementary_exams')
@nav_item(title="Supplementary Exam Marks", icon="simple-icon-graduation",
          url_name="academics:supplementary_exam_management",
          permission='academics.manage_supplementary_exams', group='admin_management', order=110)
def supplementary_exam_management_view(request):
    # Get all records where the student has failed and is eligible
    eligible_students = StudentSubjectStatus.objects.filter(
        status='FAILED'
    ).select_related(
        'student__profile', 'subject__subject', 'subject__course'
    ).order_by('subject__subject', 'student__username')

    context = {
        'page_title': 'Supplementary Exam Management',
        'eligible_students': eligible_students,
    }
    return render(request, 'academics/supplementary_management.html', context)


@login_required
@permission_required('academics.change_mark')  # Or a more specific permission
def enter_supplementary_marks_view(request, status_pk):
    status_record = get_object_or_404(StudentSubjectStatus, pk=status_pk)

    # We need to find the specific mark record to update.
    # This assumes the supplementary exam replaces the 'External' or main exam mark.
    # You may need to adjust the filter based on your specific Criterion names.
    try:
        mark_to_update = Mark.objects.get(
            student=status_record.student,
            subject=status_record.subject,
            criterion__name__icontains='external'  # Or 'Final Exam', etc.
        )
    except Mark.DoesNotExist:
        messages.error(request, "Could not find the original mark record to update.")
        return redirect('academics:supplementary_exam_management')

    if request.method == 'POST':
        form = SupplementaryMarkForm(request.POST)
        if form.is_valid():
            # --- Update the original mark ---
            mark_to_update.marks_obtained = form.cleaned_data['marks_obtained']
            mark_to_update.save()

            # --- Recalculate the total percentage ---
            settings = AttendanceSettings.load()
            all_marks_for_subject = Mark.objects.filter(student=status_record.student, subject=status_record.subject)
            total_obtained = sum(m.marks_obtained for m in all_marks_for_subject)
            total_max = sum(m.criterion.max_marks for m in all_marks_for_subject)
            percentage = (total_obtained / total_max * 100) if total_max > 0 else 0

            # --- Update the final status ---
            if percentage >= settings.passing_percentage:
                status_record.status = 'PASSED_SUPPLEMENTARY'
            else:
                status_record.status = 'FAILED_SUPPLEMENTARY'
            status_record.save()

            messages.success(request, f"Marks updated successfully for {status_record.student.username}.")
            return redirect('academics:supplementary_exam_management')
    else:
        form = SupplementaryMarkForm()

    context = {
        'page_title': 'Enter Supplementary Marks',
        'form': form,
        'status_record': status_record
    }
    return render(request, 'academics/supplementary_mark_form.html', context)
